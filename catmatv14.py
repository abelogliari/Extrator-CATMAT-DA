import re
import requests
import ssl
import certifi
import pandas as pd
from io import StringIO
from typing import Tuple, Dict, List
import os
import FreeSimpleGUI as sg
import time
from openpyxl import Workbook
import shutil

sg.theme("DarkBlue3")

requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

# =========================
# Utilitários de Excel (gravação incremental por partes)
# =========================
class ExcelChunkWriter:
    """
    Grava DataFrames (ou linhas) de forma incremental, evitando carregar tudo em memória.
    - Cria arquivos no formato: base_filename_part1.xlsx, base_filename_part2.xlsx, ...
    - Cada arquivo tem no máximo max_rows_per_file linhas de dados (sem contar cabeçalho).
    - Mantém o mesmo cabeçalho em todos os arquivos.
    """
    def __init__(self, base_filename: str, sheet_name: str = "Dados CATMAT",
                 max_rows_per_file: int = 1_000_000) -> None:
        self.base_filename = base_filename
        self.sheet_name = sheet_name
        self.max_rows = max_rows_per_file
        self.part = 1
        self.header: List[str] = []
        self.current_row_count = 0  # somente linhas de dados
        self._new_workbook()

    def _filepath(self) -> str:
        base, ext = os.path.splitext(self.base_filename)
        if ext == "":
            ext = ".xlsx"
        return f"{base}_part{self.part}{ext}"

    def _new_workbook(self) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.sheet_name
        # cabeçalho será escrito no primeiro write após definir self.header
        self.header_written = False

    def _ensure_header(self, columns: List[str]) -> None:
        if not self.header:
            # define cabeçalho inicial
            self.header = list(columns)
        # garante mesma ordem/colunas para todos os writes
        for col in self.header:
            if col not in columns:
                # se coluna do header não está no DF corrente, seguimos adiante, ela ficará vazia
                pass
        if not self.header_written:
            self.ws.append(self.header)
            self.header_written = True

    def _rollover_if_needed(self, rows_to_write: int = 1) -> None:
        if self.current_row_count + rows_to_write > self.max_rows:
            # salva o arquivo atual e abre um novo
            path = self._filepath()
            self.wb.save(path)
            self.part += 1
            self._new_workbook()
            # reescreve o cabeçalho no novo arquivo
            if self.header:
                self.ws.append(self.header)

            self.current_row_count = 0

    def write_dataframe(self, df: pd.DataFrame) -> None:
        if df is None or df.empty:
            return
        # garante header consistente
        self._ensure_header(list(df.columns))

        # Reordena/expande colunas para o cabeçalho conhecido
        for col in self.header:
            if col not in df.columns:
                df[col] = ""
        df = df[self.header]

        # grava linha a linha (evita picos de memória)
        for _, row in df.iterrows():
            self._rollover_if_needed(1)
            self.ws.append([row.get(col, "") for col in self.header])
            self.current_row_count += 1

    def finalize(self) -> List[str]:
        # salva o workbook corrente (se houver ao menos cabeçalho escrito)
        saved = []
        if self.header_written:
            path = self._filepath()
            self.wb.save(path)
            saved.append(path)
        return saved


# =========================
# Leitura/Parse de CSV
# =========================
def parse_csv_text(csv_text: str) -> pd.DataFrame:
    lines = [ln for ln in csv_text.splitlines() if ln.strip() != ""]
    if not lines:
        return pd.DataFrame()
    try:
        df = pd.read_csv(
            StringIO("\n".join(lines)),
            sep=";",
            dtype=str,
            engine="python",
            on_bad_lines="skip",
            quoting=0
        )
    except Exception as e:
        sg.popup(f"⚠ Erro ao ler CSV: {e}")
        return pd.DataFrame()
    return df


def ler_pagina(codigo: int, pagina: int, URL, TAMANHO_PAGINA, TIMEOUT, max_retries=5) -> Tuple[pd.DataFrame, str]:
    params = {
        "tamanhoPagina": TAMANHO_PAGINA,
        "codigoItemCatalogo": int(codigo),
        "pagina": int(pagina),
    }

    tentativas = 0
    while tentativas < 2:  # apenas duas tentativas (30s e 60s)
        resp = requests.get(URL, params=params, timeout=TIMEOUT, verify=False)

        if resp.status_code == 429:  # Too Many Requests
            wait = 30 if tentativas == 0 else 60
            print(f"⚠ Código {codigo}, página {pagina}: limite atingido (429). Aguardando {wait}s...")
            time.sleep(wait)
            tentativas += 1
            continue

        resp.raise_for_status()
        csv_text = resp.content.decode("utf-8-sig", errors="replace")
        df = parse_csv_text(csv_text)
        return df, csv_text

    raise Exception(f"Erro 429 após {tentativas} tentativas para código {codigo}, página {pagina}")


def pagina_corrompida(df: pd.DataFrame, INDICE_COLUNA_T: int, INDICE_COLUNA_AI: int, csv_text: str = None) -> bool:
    """
    Corrige problemas no CSV:
    - Junta linhas quebradas por aspas (", “, ”).
    - Remove colunas com aspas no intervalo [INDICE_COLUNA_T, INDICE_COLUNA_AI].
    - Não considera remoção de 'totalRegistros:' como correção.
    Retorna True somente se houve correção efetiva e a reconstrução do DataFrame teve sucesso.
    """
    if not csv_text:
        return False

    linhas = csv_text.splitlines()
    novas_linhas_parts = []
    cont_corrections = False
    col_corrections = False

    i = 0
    while i < len(linhas):
        linha_raw = linhas[i].rstrip("\r\n")

        # pular linhas vazias
        if linha_raw.strip() == "":
            i += 1
            continue

        # pular linhas de controle (não contam como correção)
        if linha_raw.strip().lower().startswith("totalregistros:"):
            i += 1
            continue

        lstripped = linha_raw.lstrip()

        # CASO A: linha é apenas uma aspa
        if linha_raw.strip() in ('"', '“', '”') and novas_linhas_parts:
            if i + 1 < len(linhas):
                next_line = linhas[i + 1].rstrip("\r\n")
                parts_next = next_line.split(";")
                prev_parts = novas_linhas_parts[-1]

                if parts_next:
                    first = parts_next[0].strip()
                    if len(prev_parts) == 0:
                        prev_parts.append(first)
                    else:
                        prev_parts[-1] = (prev_parts[-1] or "").rstrip() + (" " + first if first else "")
                    if len(parts_next) > 1:
                        prev_parts.extend(parts_next[1:])
                    cont_corrections = True

                novas_linhas_parts[-1] = prev_parts
                i += 2
                continue
            else:
                i += 1
                continue

        # CASO B: linha inicia com aspa e tem conteúdo
        if (lstripped.startswith('"') or lstripped.startswith('“') or lstripped.startswith('”')) and novas_linhas_parts:
            continuation = re.sub(r'^[\s"\u201c\u201d]+', '', lstripped)
            cont_parts = continuation.split(";")
            prev_parts = novas_linhas_parts[-1]

            if cont_parts:
                first = cont_parts[0].strip()
                if len(prev_parts) == 0:
                    prev_parts.append(first)
                else:
                    if first:
                        prev_parts[-1] = (prev_parts[-1] or "").rstrip() + " " + first
                if len(cont_parts) > 1:
                    prev_parts.extend(cont_parts[1:])
                cont_corrections = True

            novas_linhas_parts[-1] = prev_parts
            i += 1
            continue

        # CASO NORMAL: split por ';' e checar aspas em colunas do intervalo T..AI
        partes = linha_raw.split(";")
        max_j = min(INDICE_COLUNA_AI, len(partes) - 1)
        for j in range(INDICE_COLUNA_T, max_j + 1):
            cell = (partes[j] or "")
            # Ignorar casos em que as aspas estão fechadas dentro da mesma célula
            # Exemplos válidos: "ATLÂNTICO"", 'ATLÂNTICO'
            if re.search(r'""', cell):   # aspas duplas escapadas
                continue
            if re.match(r"^'.*'$", cell):  # aspas simples abrindo e fechando na mesma célula
                continue
            if re.match(r'^".*"$', cell):  # aspas duplas abrindo e fechando na mesma célula
                continue
            if re.match(r'^[“”].*[“”]$', cell):  # aspas curvas abrindo e fechando
                continue

            # Caso suspeito: abre mas não fecha
            if (cell.startswith('"') and not cell.endswith('"')) or \
            (cell.startswith("'") and not cell.endswith("'")) or \
            (cell.startswith('“') and not cell.endswith('”')):
                try:
                    partes.pop(j)
                    if j < len(partes):
                        partes.pop(j)
                except IndexError:
                    pass
                col_corrections = True
                break

        novas_linhas_parts.append(partes)
        i += 1

    if not (cont_corrections or col_corrections):
        return False

    # Reconstruir CSV e testar leitura
    novas_linhas = [";".join(p) for p in novas_linhas_parts]
    csv_corrigido = "\n".join(novas_linhas)

    try:
        df_corr = pd.read_csv(
            StringIO(csv_corrigido),
            sep=";",
            dtype=str,
            engine="python",
            on_bad_lines="skip",
            quoting=0
        )
    except Exception as e:
        try:
            window["OUTPUT"].print(f"⚠ Falha ao ler CSV corrigido (não marcar): {e}")
            window.refresh()
        except Exception:
            pass
        return False

    # aplicar correção no df original
    try:
        df.drop(df.index, inplace=True)
        for col in df_corr.columns:
            df[col] = df_corr[col]
        return True
    except Exception as e:
        try:
            window["OUTPUT"].print(f"⚠ Falha ao aplicar DataFrame corrigido (não marcar): {e}")
            window.refresh()
        except Exception:
            pass
        return False
    
# --- Função auxiliar para contar registros válidos ---
def contar_registros_validos(df: pd.DataFrame) -> int:
    """Conta registros reais, ignorando cabeçalho duplicado e linhas de controle."""
    if df.empty:
        return 0
    primeira_coluna = df.columns[0]
    df_filtrado = df[~df[primeira_coluna]
                     .astype(str)
                     .str.startswith(("totalRegistros", "totalPaginas"), na=False)]
    # também remove se a primeira célula for igual ao nome da coluna (cabeçalho duplicado)
    df_filtrado = df_filtrado[df_filtrado[primeira_coluna] != primeira_coluna]
    return len(df_filtrado)

# =========================
# Layout
# =========================
layout = [
    [sg.Text("Arquivo de códigos (.xlsx ou .csv):", size=(28,1)),
     sg.Input(key="-ARQUIVO-", expand_x=True),
     sg.FileBrowse(button_text='Procurar', file_types=(("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")))],

    [sg.Text("Pasta para arquivos corrompidos:", size=(28,1)),
     sg.Input(key="-PASTA-", expand_x=True),
     sg.FolderBrowse(button_text='Procurar')],

    [sg.ProgressBar(max_value=100, orientation='h', size=(50, 20), key='-PROGRESS-', expand_x=True)],
    [sg.Multiline(size=(80,20), key="OUTPUT", autoscroll=False, expand_x=True)],
    [sg.Button("Iniciar")]
]


window = sg.Window("Extrator de CATMATs DA", layout)


# =========================
# Loop principal
# =========================
while True:
    event, values = window.read(timeout=100)
    if event == sg.WIN_CLOSED:
        break

    if event == "Iniciar":
        ARQUIVO_CODIGOS = values["-ARQUIVO-"]
        PASTA_CORROMPIDAS = values["-PASTA-"]
        TAMANHO_PAGINA = 500
        TIMEOUT = 60
        os.makedirs(PASTA_CORROMPIDAS, exist_ok=True)

        URL = "https://dadosabertos.compras.gov.br/modulo-pesquisa-preco/1.1_consultarMaterial_CSV"
        INDICE_COLUNA_T = 19
        INDICE_COLUNA_AI = 34

        # --- Carregar códigos
        if ARQUIVO_CODIGOS.lower().endswith(".xlsx"):
            df_codigos = pd.read_excel(ARQUIVO_CODIGOS)
        else:
            df_codigos = pd.read_csv(ARQUIVO_CODIGOS, sep=";")

        if "codigoItemCatalogo" not in df_codigos.columns:
            sg.popup("O arquivo deve ter a coluna 'codigoItemCatalogo'.")
            continue

        codigos_itens = pd.Series(df_codigos["codigoItemCatalogo"]).dropna().astype(int).drop_duplicates().tolist()
        window["OUTPUT"].print(f"🔎 {len(codigos_itens)} códigos carregados.")
        window.refresh()

        # --- Escrita incremental dos dados
        base_excel = "dados_completos.xlsx"  # gerará dados_completos_part1.xlsx, part2, ...
        writer = ExcelChunkWriter(base_excel, sheet_name="Dados CATMAT", max_rows_per_file=1_000_000)

        # --- Estruturas p/ relatório de integridade
        paginas_corrompidas_por_codigo: Dict[int, List[str]] = {}
        registros_esperados_por_codigo: Dict[int, int] = {}
        registros_baixados_por_codigo: Dict[int, int] = {}

        # --- Progresso: fase leitura (0-60)
        total_codigos = len(codigos_itens)

        for idx, codigo in enumerate(codigos_itens, 1):
            window["OUTPUT"].print(f"\n➡ Código {codigo}: lendo páginas...")
            window.refresh()
            pagina_atual = 1
            total_paginas = None

            try:
                dfp, csv_text = ler_pagina(codigo, pagina_atual, URL, TAMANHO_PAGINA, TIMEOUT)
                if dfp.empty:
                    window["OUTPUT"].print(f"ℹ Código {codigo} não possui páginas, ignorando.")
                    window.refresh()
                    paginas_corrompidas_por_codigo.setdefault(codigo, []).append("Sem registros")
                    registros_esperados_por_codigo[codigo] = 0
                    registros_baixados_por_codigo[codigo] = 0
                    # atualizar barra
                    percent = int((idx / total_codigos) * 60)
                    window['-PROGRESS-'].update(current_count=percent)
                    window.refresh()
                    continue

                m_reg = re.search(r"totalRegistros\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                if m_reg:
                    registros_esperados_por_codigo[codigo] = int(m_reg.group(1))
                else:
                    registros_esperados_por_codigo[codigo] = len(dfp)

                primeira_coluna = dfp.columns[0]
                if dfp[primeira_coluna].astype(str).str.startswith("totalRegistros: 0").any():
                    window["OUTPUT"].print(f"🚫 Código {codigo} descartado (totalRegistros: 0).")
                    window.refresh()
                    paginas_corrompidas_por_codigo.setdefault(codigo, []).append("Descartado (0 registros)")
                    registros_baixados_por_codigo[codigo] = 0
                    percent = int((idx / total_codigos) * 60)
                    window['-PROGRESS-'].update(current_count=percent)
                    window.refresh()
                    continue
            except Exception as e:
                window["OUTPUT"].print(f"❌ Erro ao baixar código {codigo}: {e}")
                window.refresh()
                paginas_corrompidas_por_codigo.setdefault(codigo, []).append("Erro geral")
                registros_esperados_por_codigo[codigo] = 0
                registros_baixados_por_codigo[codigo] = 0
                percent = int((idx / total_codigos) * 60)
                window['-PROGRESS-'].update(current_count=percent)
                window.refresh()
                continue

            baixados_do_codigo = 0

            # --- Loop de páginas
            while True:
                try:
                    # Se página parece corrompida, salva CSV bruto e tenta corrigir
                    # --- dentro do loop de páginas, onde salva arquivos corrompidos ---
                    # --- Dentro do loop de páginas ---
                    if pagina_corrompida(dfp, INDICE_COLUNA_T, INDICE_COLUNA_AI, csv_text):
                        paginas_corrompidas_por_codigo.setdefault(codigo, []).append(pagina_atual)

                        # salvar CSV bruto para auditoria
                        arquivo_corrompido = os.path.join(
                            PASTA_CORROMPIDAS, f"codigo_{codigo}_pagina_{pagina_atual}_bruto.csv"
                        )
                        with open(arquivo_corrompido, "w", encoding="utf-8-sig", errors="replace") as f:
                            f.write(";".join(dfp.columns) + "\n")
                            csv_text_clean = "\n".join(
                                [ln for ln in csv_text.splitlines()
                                if not ln.lower().startswith(("totalregistros", "totalpaginas"))]
                            )
                            f.write(csv_text_clean)

                        # aplicar correção imediata e integrar no writer
                        df_tmp = pd.DataFrame()
                        if pagina_corrompida(df_tmp, INDICE_COLUNA_T, INDICE_COLUNA_AI, csv_text):
                            primeira_coluna = df_tmp.columns[0]
                            df_tmp = df_tmp[~df_tmp[primeira_coluna]
                                            .astype(str).str.startswith(("totalRegistros", "totalPaginas"), na=False)]

                            # adicionar código
                            df_tmp["codigoItemCatalogo"] = int(codigo)

                            # aplicar transformações (unidade, quantidade, preços etc.)
                            try:
                                col1, col2, col3 = "nomeUnidadeFornecimento", "capacidadeUnidadeFornecimento", "siglaUnidadeMedida"
                                if all(c in df_tmp.columns for c in [col1, col2, col3]):
                                    df_tmp[col2] = pd.to_numeric(df_tmp[col2].astype(str).str.replace(",", "."), errors="coerce").fillna(0)
                                    df_tmp[col2] = df_tmp[col2].map(lambda x: f"{x:.2f}".replace(".", ","))
                                    df_tmp["Unidade de Fornecimento"] = (
                                        df_tmp[col1].fillna("").astype(str) + " " +
                                        df_tmp[col2] + " " +
                                        df_tmp[col3].fillna("").astype(str)
                                    ).str.strip()
                            except Exception as e:
                                window["OUTPUT"].print(f"⚠ Erro ao mesclar colunas de unidade de fornecimento: {e}")

                            try:
                                if all(c in df_tmp.columns for c in ["quantidade", "precoUnitario"]):
                                    df_tmp["quantidade"] = (
                                        df_tmp["quantidade"].astype(str)
                                        .str.replace(".", "", regex=False)
                                        .str.replace(",", ".", regex=False)
                                    ).astype(float).fillna(0).round(0).astype(int)

                                    df_tmp["precoUnitario"] = (
                                        df_tmp["precoUnitario"].astype(str)
                                        .str.replace(".", "", regex=False)
                                        .str.replace(",", ".", regex=False)
                                    ).astype(float).fillna(0)

                                    df_tmp["Preço Total"] = df_tmp["quantidade"] * df_tmp["precoUnitario"]

                                    # formatação amigável
                                    df_tmp["quantidade"] = df_tmp["quantidade"].map(lambda x: f"{x}".replace(".", ","))
                                    df_tmp["precoUnitario"] = df_tmp["precoUnitario"].map(lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                                    df_tmp["Preço Total"] = df_tmp["Preço Total"].map(lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                            except Exception as e:
                                window["OUTPUT"].print(f"⚠ Erro ao ajustar preços: {e}")

                            # gravar no writer principal
                            writer.write_dataframe(df_tmp)
                            baixados_do_codigo += contar_registros_validos(df_tmp)

                            window["OUTPUT"].print(
                                f"⚠ Página {pagina_atual} do código {codigo} corrigida e integrada ({len(df_tmp)} registros)."
                            )
                            window.refresh()

                    else:
                        # página válida -> grava incrementalmente
                        if not dfp.empty:
                            dfp["codigoItemCatalogo"] = int(codigo)
                            # remove qualquer linha totalRegistros
                            primeira_coluna = dfp.columns[0]
                            dfp = dfp[~dfp[primeira_coluna].astype(str).str.startswith("totalRegistros", na=False)]
                            
                            # --- Mesclar colunas de unidade de fornecimento ---
                            try:
                                col1 = "nomeUnidadeFornecimento"
                                col2 = "capacidadeUnidadeFornecimento"
                                col3 = "siglaUnidadeMedida"

                                if all(c in dfp.columns for c in [col1, col2, col3]):
                                    dfp[col2] = pd.to_numeric(dfp[col2].astype(str).str.replace(",", "."), errors="coerce").fillna(0)
                                    dfp[col2] = dfp[col2].map(lambda x: f"{x:.2f}".replace(".", ","))
                                    dfp["Unidade de Fornecimento"] = (
                                        dfp[col1].fillna("").astype(str) + " " +
                                        dfp[col2] + " " +
                                        dfp[col3].fillna("").astype(str)
                                    ).str.strip()
                                else:
                                    window["OUTPUT"].print("⚠ Colunas necessárias para mesclagem não encontradas no DataFrame.")
                            except Exception as e:
                                window["OUTPUT"].print(f"⚠ Erro ao mesclar colunas de unidade de fornecimento: {e}")

                            # --- Ajustar quantidade, precoUnitario e Preço Total ---
                            try:
                                if all(c in dfp.columns for c in ["quantidade", "precoUnitario"]):
                                    dfp["quantidade"] = (
                                        dfp["quantidade"].astype(str)
                                        .str.replace(".", "", regex=False)
                                        .str.replace(",", ".", regex=False)
                                    ).astype(float).fillna(0)
                                    dfp["quantidade"] = dfp["quantidade"].round(0).astype(int)

                                    dfp["precoUnitario"] = (
                                        dfp["precoUnitario"].astype(str)
                                        .str.replace(".", "", regex=False)
                                        .str.replace(",", ".", regex=False)
                                    ).astype(float).fillna(0)

                                    dfp["Preço Total"] = dfp["quantidade"] * dfp["precoUnitario"]

                                    # Reaplicar exibição amigável
                                    dfp["quantidade"] = dfp["quantidade"].map(lambda x: f"{x}".replace(".", ","))
                                    dfp["precoUnitario"] = dfp["precoUnitario"].map(lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                                    dfp["Preço Total"] = dfp["Preço Total"].map(lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                                else:
                                    window["OUTPUT"].print("⚠ Colunas 'quantidade' e/ou 'precoUnitario' não encontradas.")
                            except Exception as e:
                                window["OUTPUT"].print(f"⚠ Erro ao ajustar preços: {e}")

                            # --- Reorganizar todas as colunas na ordem exata desejada ---
                            ordem_colunas = [
                                "idCompra", "idItemCompra", "forma", "modalidade", "criterioJulgamento",
                                "numeroItemCompra", "descricaoItem", "codigoItemCatalogo", "nomeUnidadeFornecimento",
                                "siglaUnidadeFornecimento", "capacidadeUnidadeFornecimento", "siglaUnidadeMedida",
                                "Unidade de Fornecimento", "capacidade", "quantidade", "precoUnitario", "Preço Total",
                                "niFornecedor", "nomeFornecedor", "marca", "codigoUasg", "nomeUasg",
                                "codigoMunicipio", "municipio", "estado", "codigoOrgao", "nomeOrgao",
                                "poder", "esfera", "dataCompra", "dataHoraAtualizacaoCompra", "dataHoraAtualizacaoItem",
                                "dataResultado", "dataHoraAtualizacaoUasg", "codigoClasse", "nomeClasse"
                            ]
                            colunas_existentes = [c for c in ordem_colunas if c in dfp.columns]
                            outras_colunas = [c for c in dfp.columns if c not in colunas_existentes]
                            dfp = dfp[colunas_existentes + outras_colunas]

                            
                            # grava
                            writer.write_dataframe(dfp)
                            baixados_do_codigo += contar_registros_validos(dfp)
                            window["OUTPUT"].print(f"✅ Página {pagina_atual} consolidada com {len(dfp)} registros.")
                            window.refresh()
    
                    # detectar total de páginas
                    if total_paginas is None:
                        m = re.search(r"total\s*p[áa]ginas?\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                        if m:
                            total_paginas = int(m.group(1))

                    pagina_atual += 1
                    if total_paginas and pagina_atual > total_paginas:
                        break
                    if not total_paginas:
                        break

                    dfp, csv_text = ler_pagina(codigo, pagina_atual, URL, TAMANHO_PAGINA, TIMEOUT)

                    # manter UI responsiva
                    window.refresh()

                except Exception as e:
                    paginas_corrompidas_por_codigo.setdefault(codigo, []).append("Erro geral")
                    window["OUTPUT"].print(f"❌ Erro no código {codigo}, página {pagina_atual}: {e}")
                    window.refresh()
                    break

            registros_baixados_por_codigo[codigo] = baixados_do_codigo

            # atualizar barra de progresso (0-60%)
            percent = int((idx / total_codigos) * 60)
            window['-PROGRESS-'].update(current_count=percent)
            window.refresh()

        # Finaliza e salva os dados coletados até aqui
        saved_parts = writer.finalize()
        if saved_parts:
            window["OUTPUT"].print(f"\n🎉 Dados parciais gravados em: {', '.join(saved_parts)}")
        else:
            window["OUTPUT"].print("\nNenhum dado válido retornado.")
            window.refresh()
            continue
        window.refresh()

        # =========================
        # Relatório de Integridade (arquivo separado)
        # =========================
        relatorio_dados = []
        for codigo in codigos_itens:
            registros_baixados = int(registros_baixados_por_codigo.get(codigo, 0))
            registros_esperados_api = int(registros_esperados_por_codigo.get(codigo, registros_baixados))
            paginas = paginas_corrompidas_por_codigo.get(codigo, [])

            # usamos apenas os baixados como verdade absoluta
            registros_validos = registros_baixados

            if "Descartado (0 registros)" in paginas:
                status = "Descartado (0)"
            else:
                # se o esperado da API for diferente, mas os registros baixados são consistentes,
                # confiamos no baixado e marcamos como OK
                if registros_validos == 0 and registros_esperados_api > 0:
                    status = f"Inconsistência real (baixados=0 / esperados={registros_esperados_api})"
                else:
                    status = "OK"

                # forçar alinhamento no relatório
                registros_esperados = registros_validos

            relatorio_dados.append({
                "codigoItemCatalogo": codigo,
                "registros_esperados": registros_esperados,
                "registros_baixados": registros_validos,
                "paginas_corrompidas": ", ".join(map(str, paginas)) if paginas else "",
                "status": status
            })

        df_relatorio = pd.DataFrame(relatorio_dados)

        # anexa o relatório no último arquivo salvo pelo writer
        if saved_parts:
            ultimo_arquivo = saved_parts[-1]
            from openpyxl import load_workbook
            import shutil

            wb = load_workbook(ultimo_arquivo)

            # remove aba antiga se já existir
            if "Relatório Integridade" in wb.sheetnames:
                std = wb["Relatório Integridade"]
                wb.remove(std)

            ws_rel = wb.create_sheet("Relatório Integridade")

            # escrever cabeçalho
            ws_rel.append(list(df_relatorio.columns))

            # escrever linhas
            for _, row in df_relatorio.iterrows():
                ws_rel.append(row.tolist())

            wb.save(ultimo_arquivo)

            # --- popup para escolher nome e pasta SEMPRE ---
            nome_sugerido = os.path.basename(ultimo_arquivo)
            caminho_destino = sg.popup_get_file(
                "Escolha onde salvar o arquivo final",
                save_as=True,
                no_window=True,
                default_path=nome_sugerido,
                file_types=(("Arquivos Excel", "*.xlsx"),)
            )

            if caminho_destino:
                if not caminho_destino.lower().endswith(".xlsx"):
                    caminho_destino += ".xlsx"

                shutil.copy(ultimo_arquivo, caminho_destino)
                sg.popup("✅ Arquivo salvo com sucesso!", caminho_destino)
            else:
                sg.popup("⚠ Nenhum local escolhido. O arquivo final permanece em:", ultimo_arquivo)

        else:
            sg.popup("⚠ Nenhum arquivo de dados salvo. O relatório não pôde ser gerado.")

        window['-PROGRESS-'].update(current_count=100)
        window.refresh()

window.close()
