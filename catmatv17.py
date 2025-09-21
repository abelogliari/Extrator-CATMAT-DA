import re
import requests
import pandas as pd
from io import StringIO
from typing import Tuple, Dict, List, Optional
import os
import FreeSimpleGUI as sg
import time
from openpyxl import Workbook, load_workbook
import shutil

# --- Funções e Classes de Lógica de Negócio (Sem alterações) ---
# Nenhuma mudança nesta seção. O "motor" do programa continua o mesmo.

requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

class ExcelChunkWriter:
    def __init__(self, base_filename: str, sheet_name: str = "Dados CATMAT", max_rows_per_file: int = 1_000_000):
        self.base_filename = base_filename
        self.sheet_name = sheet_name
        self.max_rows = max_rows_per_file
        self.part = 1
        self.header: List[str] = []
        self.current_row_count = 0
        self.files_saved = []
        self._new_workbook()

    def _filepath(self) -> str:
        base, ext = os.path.splitext(self.base_filename)
        return f"{base}_part{self.part}{ext or '.xlsx'}"

    def _new_workbook(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.sheet_name
        self.header_written = False
        self.current_row_count = 0

    def _ensure_header(self, columns: List[str]):
        if not self.header: self.header = list(columns)
        if not self.header_written:
            self.ws.append(self.header)
            self.header_written = True

    def _rollover_if_needed(self, rows_to_write: int = 1):
        if self.current_row_count + rows_to_write > self.max_rows:
            path = self._filepath()
            self.wb.save(path)
            self.files_saved.append(path)
            self.part += 1
            self._new_workbook()
            if self.header: self.ws.append(self.header)

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        self._ensure_header(list(df.columns))
        for col in self.header:
            if col not in df.columns: df[col] = ""
        df = df[self.header]
        for _, row in df.iterrows():
            self._rollover_if_needed(1)
            self.ws.append([row.get(col, "") for col in self.header])
            self.current_row_count += 1

    def finalize(self) -> List[str]:
        if self.header_written and self.current_row_count > 0:
            path = self._filepath()
            self.wb.save(path)
            if path not in self.files_saved: self.files_saved.append(path)
        return self.files_saved

def parse_csv_text(csv_text: str) -> pd.DataFrame:
    lines = [ln for ln in csv_text.splitlines() if ln.strip()]
    if not lines: return pd.DataFrame()
    try:
        return pd.read_csv(StringIO("\n".join(lines)), sep=";", dtype=str, engine="python", on_bad_lines="skip", quoting=0)
    except Exception as e:
        sg.popup_error(f"⚠ Erro ao ler CSV: {e}")
        return pd.DataFrame()

def ler_pagina(codigo: int, pagina: int, URL, TAMANHO_PAGINA, TIMEOUT) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    params = {"tamanhoPagina": TAMANHO_PAGINA, "codigoItemCatalogo": int(codigo), "pagina": int(pagina)}
    tentativas = 0
    while tentativas < 2:
        try:
            resp = requests.get(URL, params=params, timeout=TIMEOUT, verify=False)
            if resp.status_code == 429:
                wait = 30 if tentativas == 0 else 60
                print(f"⚠ Código {codigo}, página {pagina}: limite atingido (429). Aguardando {wait}s...")
                time.sleep(wait)
                tentativas += 1
                continue
            resp.raise_for_status()
            csv_text = resp.content.decode("utf-8-sig", errors="replace")
            df = parse_csv_text(csv_text)
            return df, csv_text
        except requests.exceptions.RequestException as e:
            raise Exception(f"Erro de rede ao buscar página {pagina} do código {codigo}: {e}")
    raise Exception(f"Erro 429 persistente para código {codigo}, página {pagina}")


def pagina_corrompida(csv_text: str, INDICE_COLUNA_T: int, INDICE_COLUNA_AI: int) -> Tuple[bool, Optional[pd.DataFrame], str]:
    # (Esta função permanece exatamente a mesma)
    if not csv_text: return False, None, csv_text
    linhas = csv_text.splitlines()
    novas_linhas_parts: List[List[str]] = []
    cont_corrections, col_corrections = False, False
    i = 0
    while i < len(linhas):
        linha_raw = linhas[i].rstrip("\r\n")
        if not linha_raw.strip() or linha_raw.strip().lower().startswith(("totalregistros:", "totalpaginas:")): i += 1; continue
        lstripped = linha_raw.lstrip()
        if linha_raw.rstrip().endswith('"') and (linha_raw.count('"') % 2 == 1):
            if i + 1 < len(linhas): linha_raw += linhas[i + 1].lstrip(); cont_corrections = True; i += 2
            else: i += 1; continue
        else:
            if lstripped in ('"', '“', '”') and novas_linhas_parts:
                if i + 1 < len(linhas):
                    next_line, parts_next = linhas[i + 1].rstrip("\r\n"), linhas[i + 1].rstrip("\r\n").split(";")
                    prev_parts = novas_linhas_parts[-1]
                    if parts_next:
                        first = parts_next[0].strip()
                        if not prev_parts: prev_parts.append(first)
                        else: prev_parts[-1] = (prev_parts[-1] or "").rstrip() + (" " + first if first else "")
                        if len(parts_next) > 1: prev_parts.extend(parts_next[1:])
                        cont_corrections = True
                    novas_linhas_parts[-1] = prev_parts; i += 2; continue
                else: i += 1; continue
            if (lstripped.startswith('"') or lstripped.startswith('“') or lstripped.startswith('”')) and novas_linhas_parts:
                continuation, cont_parts = re.sub(r'^[\s"\u201c\u201d]+', '', lstripped), re.sub(r'^[\s"\u201c\u201d]+', '', lstripped).split(";")
                prev_parts = novas_linhas_parts[-1]
                if cont_parts:
                    first = cont_parts[0].strip()
                    if not prev_parts: prev_parts.append(first)
                    else:
                        if first: prev_parts[-1] = (prev_parts[-1] or "").rstrip() + " " + first
                    if len(cont_parts) > 1: prev_parts.extend(cont_parts[1:])
                    cont_corrections = True
                novas_linhas_parts[-1] = prev_parts; i += 1; continue
            i += 1
        partes = linha_raw.split(";")
        max_j = min(len(partes) - 1, INDICE_COLUNA_AI)
        for j in range(INDICE_COLUNA_T, max_j + 1):
            cell = partes[j] or ""
            if any([re.search(r'""', cell), re.match(r"^'.*'$", cell), re.match(r'^".*"$', cell), re.match(r'^[“”].*[“”]$', cell)]): continue
            if (cell.startswith('"') and not cell.endswith('"')) or (cell.startswith("'") and not cell.endswith("'")) or (cell.startswith('“') and not cell.endswith('”')):
                try:
                    partes.pop(j);
                    if j < len(partes): partes.pop(j)
                except IndexError: pass
                col_corrections = True; break
        novas_linhas_parts.append(partes)
    if not (cont_corrections or col_corrections): return False, None, csv_text
    csv_corrigido = "\n".join([";".join(p) for p in novas_linhas_parts])
    try:
        df_corr = pd.read_csv(StringIO(csv_corrigido), sep=";", dtype=str, engine="python", on_bad_lines="skip", quoting=0)
        return True, df_corr, csv_corrigido
    except Exception as e:
        print(f"[pagina_corrompida] Erro ao reler CSV corrigido: {e}"); return True, None, csv_corrigido

def contar_registros_validos(df: pd.DataFrame) -> int:
    if df.empty: return 0
    primeira_coluna = df.columns[0]
    df_filtrado = df[~df[primeira_coluna].astype(str).str.contains("totalRegistros|totalPaginas", case=False, na=False)]
    df_filtrado = df_filtrado[df_filtrado[primeira_coluna] != primeira_coluna]
    return len(df_filtrado)


# --- NOVA ESTRUTURA DA INTERFACE GRÁFICA (UI) APRIMORADA ---

sg.theme('SystemDefaultForReal')

layout_config = [
    [sg.Text("Arquivo de Códigos:", size=(20,1)), sg.Input(key="-ARQUIVO-", enable_events=True, expand_x=True), sg.FileBrowse(button_text='Procurar', file_types=(("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")))],
    [sg.Text("Pasta para Corrompidos:", size=(20,1), tooltip="Pasta onde serão salvos os arquivos CSV que apresentaram erros."), sg.Input(key="-PASTA-", enable_events=True, expand_x=True), sg.FolderBrowse(button_text='Procurar')],
]

layout_stats = [
    [sg.Text("Códigos Processados:", size=(20,1)), sg.Text("0 / 0", key="-CONT_PROCESSADOS-", font=("Helvetica", 10, "bold"))],
    [sg.Text("Registros Consolidados:", size=(20,1)), sg.Text("0", key="-CONT_REGISTROS-", font=("Helvetica", 10, "bold"))],
    [sg.Text("Páginas Corrigidas:", size=(20,1)), sg.Text("0", key="-CONT_CORRIGIDAS-", font=("Helvetica", 10, "bold"), text_color="orange")],
    [sg.Text("Códigos sem Dados:", size=(20,1)), sg.Text("0", key="-CONT_VAZIOS-", font=("Helvetica", 10, "bold"), text_color="#FF6347")],
]

layout_execucao = [
    [sg.Text("Status: Ocioso", key='-STATUS-', expand_x=True, font=("Helvetica", 10, "italic"))],
    # ALTERAÇÃO: Adicionado o sg.Text para o percentual
    [sg.ProgressBar(max_value=1000, orientation='h', size=(50, 20), key='-PROGRESS-', expand_x=True),
     sg.Text('0%', size=(5,1), key='-PERCENT-', font=("Helvetica", 10, "bold"))],
    [sg.Multiline(size=(80,20), key="-OUTPUT-", autoscroll=True, expand_x=True, expand_y=True, background_color='black', text_color='white')],
]

layout = [
    [sg.Frame('1. Configurações de Entrada', layout_config, expand_x=True)],
    [sg.Frame('2. Resumo da Execução', layout_stats, expand_x=True)],
    [sg.Frame('3. Log e Progresso', layout_execucao, expand_x=True, expand_y=True)],
    [sg.Button("Iniciar Extração", key="-START-", disabled=True), sg.Button("Cancelar", key="-CANCEL-", disabled=True), sg.Button("Salvar Log", key="-SAVE_LOG-", disabled=True)]
]

window = sg.Window("Extrator de CATMATs Pro", layout, resizable=True, finalize=True)
window.set_min_size((700, 600))

# --- LOOP DE EVENTOS E LÓGICA DE CONTROLE APRIMORADOS ---

processing = False
codes_iterator = None
writer = None
codigos_itens, paginas_corrompidas_por_codigo, registros_esperados_por_codigo, registros_baixados_por_codigo = [], {}, {}, {}
total_registros_baixados, paginas_corrigidas_count, codigos_vazios_count = 0, 0, 0

while True:
    event, values = window.read(timeout=100)

    if event == sg.WIN_CLOSED:
        break

    if not processing:
        if values["-ARQUIVO-"] and values["-PASTA-"]: window["-START-"].update(disabled=False)
        else: window["-START-"].update(disabled=True)

    if event == "-START-" and not processing:
        processing = True
        window["-START-"].update(disabled=True); window["-CANCEL-"].update(disabled=False); window["-SAVE_LOG-"].update(disabled=True)
        window["-OUTPUT-"].update(""); window["-PROGRESS-"].update(0); window["-PERCENT-"].update('0%')
        
        window["-CONT_PROCESSADOS-"].update("0 / 0"); window["-CONT_REGISTROS-"].update("0")
        window["-CONT_CORRIGIDAS-"].update("0"); window["-CONT_VAZIOS-"].update("0")

        ARQUIVO_CODIGOS, PASTA_CORROMPIDAS = values["-ARQUIVO-"], values["-PASTA-"]
        TAMANHO_PAGINA, TIMEOUT = 500, 60
        os.makedirs(PASTA_CORROMPIDAS, exist_ok=True)
        URL = "https://dadosabertos.compras.gov.br/modulo-pesquisa-preco/1.1_consultarMaterial_CSV"
        INDICE_COLUNA_T, INDICE_COLUNA_AI = 19, 34

        try:
            df_codigos = pd.read_excel(ARQUIVO_CODIGOS) if ARQUIVO_CODIGOS.lower().endswith(".xlsx") else pd.read_csv(ARQUIVO_CODIGOS, sep=";")
            if "codigoItemCatalogo" not in df_codigos.columns:
                sg.popup_error("Erro: O arquivo de entrada deve ter a coluna 'codigoItemCatalogo'.")
                processing = False
                continue
            codigos_itens = pd.Series(df_codigos["codigoItemCatalogo"]).dropna().astype(int).drop_duplicates().tolist()
            window["-OUTPUT-"].print(f"🔎 {len(codigos_itens)} códigos únicos carregados.", text_color='lightblue')
            
            writer = ExcelChunkWriter("dados_completos.xlsx")
            paginas_corrompidas_por_codigo, registros_esperados_por_codigo, registros_baixados_por_codigo = {}, {}, {}
            total_registros_baixados, paginas_corrigidas_count, codigos_vazios_count = 0, 0, 0
            
            codes_iterator = iter(enumerate(codigos_itens, 1))
            window["-STATUS-"].update("Status: Processando...")
            window["-CONT_PROCESSADOS-"].update(f"0 / {len(codigos_itens)}")

        except Exception as e:
            sg.popup_error(f"Ocorreu um erro ao ler o arquivo de códigos:\n{e}")
            processing = False
            window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True)
            window["-STATUS-"].update("Status: Erro na inicialização")

    if event == "-CANCEL-" and processing:
        processing = False
        window["-STATUS-"].update("Status: Cancelando..."); window["-OUTPUT-"].print("\n🛑 Processo cancelado. Finalizando arquivos...", text_color='yellow')
        if writer:
            saved_parts = writer.finalize()
            if saved_parts: window["-OUTPUT-"].print(f"💾 Dados parciais salvos em: {', '.join(saved_parts)}", text_color='yellow')
        codes_iterator = None
        window["-PROGRESS-"].update(0); window["-PERCENT-"].update('0%');
        window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True)
        window["-SAVE_LOG-"].update(disabled=False); window["-STATUS-"].update("Status: Cancelado")

    if event == "-SAVE_LOG-":
        filepath = sg.popup_get_file("Salvar Log de Execução", save_as=True, no_window=True, default_extension=".txt", file_types=(("Text Files", "*.txt"),))
        if filepath:
            try:
                with open(filepath, 'w', encoding='utf-8') as f: f.write(values["-OUTPUT-"])
                sg.popup_quick("Log salvo com sucesso!")
            except Exception as e: sg.popup_error(f"Não foi possível salvar o log.\nErro: {e}")

    if processing:
        try:
            idx, codigo = next(codes_iterator)
            total_codigos = len(codigos_itens)
            window["-STATUS-"].update(f"Status: Processando código {codigo} ({idx}/{total_codigos})")
            window["-CONT_PROCESSADOS-"].update(f"{idx} / {total_codigos}")
            
            pagina_atual, total_paginas, baixados_do_codigo = 1, None, 0
            try:
                dfp, csv_text = ler_pagina(codigo, pagina_atual, URL, TAMANHO_PAGINA, TIMEOUT)
                if dfp is None or dfp.empty or csv_text is None:
                    window["-OUTPUT-"].print(f"ℹ️ Código {codigo}: Nenhum registro encontrado.", text_color='lightblue'); codigos_vazios_count += 1; window["-CONT_VAZIOS-"].update(codigos_vazios_count)
                    raise StopIteration
                
                m_reg = re.search(r"totalRegistros\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                registros_esperados = int(m_reg.group(1)) if m_reg else 0
                registros_esperados_por_codigo[codigo] = registros_esperados
                if registros_esperados == 0:
                     window["-OUTPUT-"].print(f"ℹ️ Código {codigo}: API informa 0 registros.", text_color='lightblue'); codigos_vazios_count += 1; window["-CONT_VAZIOS-"].update(codigos_vazios_count)
                     raise StopIteration
                
                while True:
                    is_corrupt, df_corrigido, _ = pagina_corrompida(csv_text, INDICE_COLUNA_T, INDICE_COLUNA_AI)
                    df_final_pagina = None
                    if is_corrupt:
                        paginas_corrompidas_por_codigo.setdefault(codigo, []).append(str(pagina_atual))
                        if df_corrigido is not None and not df_corrigido.empty:
                            window["-OUTPUT-"].print(f"⚠️ Cód {codigo}, Pág {pagina_atual}: Corrigida.", text_color='orange'); df_final_pagina = df_corrigido
                            paginas_corrigidas_count += 1; window["-CONT_CORRIGIDAS-"].update(paginas_corrigidas_count)
                        else: window["-OUTPUT-"].print(f"❌ Cód {codigo}, Pág {pagina_atual}: Falha na correção.", text_color='#FF6347')
                    else:
                        df_final_pagina = dfp
                    
                    if df_final_pagina is not None and not df_final_pagina.empty:
                        df_final_pagina["codigoItemCatalogo"] = str(codigo)
                        registros_na_pagina = contar_registros_validos(df_final_pagina)
                        baixados_do_codigo += registros_na_pagina
                        total_registros_baixados += registros_na_pagina
                        window["-CONT_REGISTROS-"].update(f"{total_registros_baixados:,}".replace(",", "."))
                        writer.write_dataframe(df_final_pagina)
                    
                    if total_paginas is None:
                        m = re.search(r"total\s*p[áa]ginas?\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                        total_paginas = int(m.group(1)) if m else 1
                    
                    pagina_atual += 1
                    if pagina_atual > total_paginas: break
                    dfp, csv_text = ler_pagina(codigo, pagina_atual, URL, TAMANHO_PAGINA, TIMEOUT)
                    if dfp is None or csv_text is None: break

            except StopIteration: pass
            except Exception as e: window["-OUTPUT-"].print(f"❌ Erro crítico no código {codigo}: {e}", text_color='#FF6347')
            registros_baixados_por_codigo[codigo] = baixados_do_codigo
            
            # ALTERAÇÃO: Atualiza a barra e o percentual juntos
            percent_complete = int((idx / total_codigos) * 100)
            window['-PROGRESS-'].update(percent_complete * 10)
            window['-PERCENT-'].update(f'{percent_complete}%')

        except StopIteration:
            # ALTERAÇÃO: Atinge 100% assim que o loop termina
            window['-PROGRESS-'].update(1000); window['-PERCENT-'].update('100%')
            window["-STATUS-"].update("Status: Processamento concluído! Gerando relatório...")
            window.refresh() # Força a UI a atualizar para 100%
            
            processing = False
            window["-OUTPUT-"].print("\n🎉 Download concluído! Gerando arquivo final...", text_color='lightblue')

            saved_parts = writer.finalize()
            if not saved_parts:
                sg.popup("Nenhum dado válido foi baixado. O relatório não será gerado.")
                window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True); window["-SAVE_LOG-"].update(disabled=False); window["-STATUS-"].update("Status: Ocioso")
                continue
            
            window["-OUTPUT-"].print(f"💾 Arquivos de dados salvos em: {', '.join(saved_parts)}", text_color='lightblue')
            
            ultimo_arquivo = saved_parts[-1]
            wb = load_workbook(ultimo_arquivo)
            if "Relatório Integridade" in wb.sheetnames: wb.remove(wb["Relatório Integridade"])
            ws_rel = wb.create_sheet("Relatório Integridade")
            ws_rel.append(["codigoItemCatalogo", "registros_esperados_api", "registros_baixados_reais", "paginas_com_problemas", "status"])
            for c in codigos_itens:
                baixados = int(registros_baixados_por_codigo.get(c, 0))
                esperados = int(registros_esperados_por_codigo.get(c, 0))
                paginas = paginas_corrompidas_por_codigo.get(c, [])
                status = "OK"
                if esperados > 0 and baixados < esperados: status = f"Inconsistência (Baixados: {baixados} / Esperados: {esperados})"
                elif esperados == 0 and baixados > 0: status = f"OK (API reportou 0, mas {baixados} foram encontrados)"
                ws_rel.append([c, esperados, baixados, ", ".join(map(str, paginas)), status])
            wb.save(ultimo_arquivo)
            
            window["-OUTPUT-"].print("📊 Relatório de integridade adicionado.", text_color='lightblue')
            
            resumo_final = f"""
✅ Processo Concluído!

-------------------------------------------
  - Códigos Processados: {len(codigos_itens)}
  - Registros Consolidados: {f"{total_registros_baixados:,}".replace(",", ".")}
  - Páginas Corrigidas: {paginas_corrigidas_count}
  - Códigos sem Registros: {codigos_vazios_count}
-------------------------------------------
            """
            sg.popup('Resumo da Extração', resumo_final)
            
            caminho_destino = sg.popup_get_file("Escolha onde salvar o arquivo final com o relatório", save_as=True, no_window=True, default_path=os.path.basename(ultimo_arquivo), file_types=(("Arquivos Excel", "*.xlsx"),))
            if caminho_destino:
                if not caminho_destino.lower().endswith(".xlsx"): caminho_destino += ".xlsx"
                shutil.copy(ultimo_arquivo, caminho_destino); sg.popup("✅ Sucesso!", f"Arquivo final salvo em:\n{caminho_destino}")
            else: sg.popup("⚠ Atenção", f"Nenhum local escolhido. O arquivo final permanece em:\n{ultimo_arquivo}")
            
            window["-STATUS-"].update("Status: Concluído!"); window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True); window["-SAVE_LOG-"].update(disabled=False)

window.close()