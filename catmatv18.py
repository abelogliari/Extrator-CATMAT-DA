import re
import requests
import pandas as pd
from io import StringIO
from typing import Tuple, Dict, List, Optional
import os
import FreeSimpleGUI as sg
import time
from openpyxl import Workbook, load_workbook
import shutil
import json
import threading

# --- Funções e Classes de Lógica de Negócio ---

requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

class ExcelChunkWriter:
    def __init__(self, base_filename: str, sheet_name: str = "Dados CATMAT", max_rows_per_file: int = 1_000_000):
        self.base_filename = base_filename
        self.sheet_name = sheet_name
        self.max_rows = max_rows_per_file
        self.part = 1
        self.header: List[str] = []
        self.current_row_count = 0
        self.files_saved = []
        self._new_workbook()

    def _filepath(self) -> str:
        base, ext = os.path.splitext(self.base_filename)
        return f"{base}_part{self.part}{ext or '.xlsx'}"

    def _new_workbook(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.sheet_name
        self.header_written = False
        self.current_row_count = 0

    def _ensure_header(self, columns: List[str]):
        if not self.header: self.header = list(columns)
        if not self.header_written:
            self.ws.append(self.header)
            self.header_written = True

    def _rollover_if_needed(self, rows_to_write: int = 1):
        if self.current_row_count + rows_to_write > self.max_rows:
            path = self._filepath()
            self.wb.save(path)
            self.files_saved.append(path)
            self.part += 1
            self._new_workbook()
            if self.header: self.ws.append(self.header)

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        self._ensure_header(list(df.columns))
        # Garante que o dataframe a ser escrito tenha todas as colunas do header
        for col in self.header:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[self.header] # Ordena o df de acordo com o header
        
        for _, row in df.iterrows():
            self._rollover_if_needed(1)
            # Converte valores para tipos nativos do Python para evitar problemas com openpyxl
            self.ws.append([
                None if pd.isna(value) else value for value in row
            ])
            self.current_row_count += 1

    def finalize(self) -> List[str]:
        if self.header_written and self.current_row_count > 0:
            path = self._filepath()
            self.wb.save(path)
            if path not in self.files_saved: self.files_saved.append(path)
        return self.files_saved


def parse_csv_text(csv_text: str) -> pd.DataFrame:
    lines = [ln for ln in csv_text.splitlines() if ln.strip()]
    if not lines: return pd.DataFrame()
    try:
        return pd.read_csv(StringIO("\n".join(lines)), sep=";", dtype=str, engine="python", on_bad_lines="skip", quoting=0)
    except Exception as e:
        sg.popup_error(f"⚠ Erro ao ler CSV: {e}")
        return pd.DataFrame()


def ler_pagina_catmat(codigo: int, pagina: int, URL_BASE, TAMANHO_PAGINA, TIMEOUT) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    URL = f"{URL_BASE}/modulo-pesquisa-preco/1.1_consultarMaterial_CSV"
    params = {"tamanhoPagina": TAMANHO_PAGINA, "codigoItemCatalogo": int(codigo), "pagina": int(pagina)}
    tentativas = 0
    while tentativas < 2:
        try:
            resp = requests.get(URL, params=params, timeout=TIMEOUT, verify=False)
            if resp.status_code == 429:
                wait = 30 if tentativas == 0 else 60; time.sleep(wait); tentativas += 1; continue
            resp.raise_for_status()
            csv_text = resp.content.decode("utf-8-sig", errors="replace")
            return parse_csv_text(csv_text), csv_text
        except requests.exceptions.RequestException as e:
            raise Exception(f"Erro de rede ao buscar CATMAT {codigo}: {e}")
    raise Exception(f"Erro 429 persistente para CATMAT {codigo}")


def buscar_pdms_por_classe(codigo_classe: int, URL_BASE: str, TIMEOUT: int) -> Optional[Tuple[pd.DataFrame, int]]:
    URL = f"{URL_BASE}/modulo-material/3_consultarPdmMaterial"
    all_pdms = []
    pagina_atual = 1
    total_paginas = 1
    total_registros_api = 0

    while pagina_atual <= total_paginas:
        params = {"codigoClasse": codigo_classe, "pagina": pagina_atual, "bps": "false"}
        try:
            resp = requests.get(URL, params=params, timeout=TIMEOUT, verify=False)
            resp.raise_for_status()
            data = resp.json()

            if "resultado" in data:
                all_pdms.extend(data["resultado"])

            if pagina_atual == 1:
                total_paginas = data.get("totalPaginas", 1)
                total_registros_api = data.get("totalRegistros", 0)

            pagina_atual += 1
            time.sleep(0.5)

        except requests.exceptions.RequestException as e:
            sg.popup_error(f"Erro de rede ao buscar PDMs da classe {codigo_classe} (página {pagina_atual}):\n{e}")
            return None
        except json.JSONDecodeError:
            sg.popup_error(f"Falha ao decodificar a resposta JSON da API para a classe {codigo_classe}.")
            return None

    if not all_pdms:
        return None

    df = pd.DataFrame(all_pdms)
    return df, total_registros_api


def buscar_catmats_por_pdm(codigos_pdm: List[int], URL_BASE: str, TIMEOUT: int, window: sg.Window) -> Optional[pd.DataFrame]:
    global cancelar_busca_catmat
    URL = f"{URL_BASE}/modulo-material/4_consultarItemMaterial"
    all_catmats = []
    total_pdms = len(codigos_pdm)

    for i, pdm_code in enumerate(codigos_pdm):
        if cancelar_busca_catmat:
            window['-STATUS_EXPLORADOR-'].update(f"Busca cancelada pelo usuário.")
            break

        pagina_atual = 1
        total_paginas = 1
        tentativas = 0
        
        window['-STATUS_EXPLORADOR-'].update(f"Buscando CATMATs do PDM {pdm_code} ({i+1}/{total_pdms})...")
        window.refresh()

        while True: 
            try:
                while pagina_atual <= total_paginas:
                    if cancelar_busca_catmat: raise InterruptedError
                    
                    params = {"codigoPdm": pdm_code, "pagina": pagina_atual, "tamanhoPagina": 500, "bps": "false"}
                    resp = requests.get(URL, params=params, timeout=TIMEOUT, verify=False)
                    
                    if resp.status_code == 429:
                        raise requests.exceptions.HTTPError("429 Client Error: Too Many Requests")

                    resp.raise_for_status()
                    data = resp.json()

                    if "resultado" in data:
                        all_catmats.extend(data["resultado"])

                    if pagina_atual == 1:
                        total_paginas = data.get("totalPaginas", 1)
                    
                    pagina_atual += 1
                    time.sleep(0.5)
                
                break 

            except requests.exceptions.HTTPError as e:
                if "429" in str(e):
                    tentativas += 1
                    if tentativas >= 2:
                        sg.popup_error(f"Erro 429 persistente para o PDM {pdm_code}. Pulando este PDM.")
                        break 
                    
                    wait = 30 if tentativas == 1 else 60
                    window['-STATUS_EXPLORADOR-'].update(f"Limite da API (429) no PDM {pdm_code}. Aguardando {wait}s...")
                    window.refresh()
                    time.sleep(wait)
                else:
                    sg.popup_error(f"Erro de rede ao buscar CATMATs do PDM {pdm_code}:\n{e}")
                    break 
            except json.JSONDecodeError:
                sg.popup_error(f"Falha ao decodificar JSON para o PDM {pdm_code}.")
                break
            except InterruptedError:
                break
        if cancelar_busca_catmat: break

    if not all_catmats:
        return None

    df = pd.DataFrame(all_catmats)
    return df


def pagina_corrompida(csv_text: str, INDICE_COLUNA_T: int, INDICE_COLUNA_AI: int) -> Tuple[bool, Optional[pd.DataFrame], str]:
    if not csv_text: return False, None, csv_text
    linhas = csv_text.splitlines()
    novas_linhas_parts: List[List[str]] = []
    cont_corrections, col_corrections = False, False
    i = 0
    while i < len(linhas):
        linha_raw = linhas[i].rstrip("\r\n")
        if not linha_raw.strip() or linha_raw.strip().lower().startswith(("totalregistros:", "totalpaginas:")): i += 1; continue
        lstripped = linha_raw.lstrip()
        if linha_raw.rstrip().endswith('"') and (linha_raw.count('"') % 2 == 1):
            if i + 1 < len(linhas): linha_raw += linhas[i + 1].lstrip(); cont_corrections = True; i += 2
            else: i += 1; continue
        else:
            if lstripped in ('"', '“', '”') and novas_linhas_parts:
                if i + 1 < len(linhas):
                    next_line, parts_next = linhas[i + 1].rstrip("\r\n"), linhas[i + 1].rstrip("\r\n").split(";")
                    prev_parts = novas_linhas_parts[-1]
                    if parts_next:
                        first = parts_next[0].strip()
                        if not prev_parts: prev_parts.append(first)
                        else: prev_parts[-1] = (prev_parts[-1] or "").rstrip() + (" " + first if first else "")
                        if len(parts_next) > 1: prev_parts.extend(parts_next[1:])
                        cont_corrections = True
                    novas_linhas_parts[-1] = prev_parts; i += 2; continue
                else: i += 1; continue
            if (lstripped.startswith('"') or lstripped.startswith('“') or lstripped.startswith('”')) and novas_linhas_parts:
                continuation, cont_parts = re.sub(r'^[\s"\u201c\u201d]+', '', lstripped), re.sub(r'^[\s"\u201c\u201d]+', '', lstripped).split(";")
                prev_parts = novas_linhas_parts[-1]
                if cont_parts:
                    first = cont_parts[0].strip()
                    if not prev_parts: prev_parts.append(first)
                    else:
                        if first: prev_parts[-1] = (prev_parts[-1] or "").rstrip() + " " + first
                    if len(cont_parts) > 1: prev_parts.extend(cont_parts[1:])
                    cont_corrections = True
                novas_linhas_parts[-1] = prev_parts; i += 1; continue
            i += 1
        partes = linha_raw.split(";")
        max_j = min(len(partes) - 1, INDICE_COLUNA_AI)
        for j in range(INDICE_COLUNA_T, max_j + 1):
            cell = partes[j] or ""
            if any([re.search(r'""', cell), re.match(r"^'.*'$", cell), re.match(r'^".*"$', cell), re.match(r'^[“”].*[“”]$', cell)]): continue
            if (cell.startswith('"') and not cell.endswith('"')) or (cell.startswith("'") and not cell.endswith("'")) or (cell.startswith('“') and not cell.endswith('”')):
                try:
                    partes.pop(j);
                    if j < len(partes): partes.pop(j)
                except IndexError: pass
                col_corrections = True; break
        novas_linhas_parts.append(partes)
    if not (cont_corrections or col_corrections): return False, None, csv_text
    csv_corrigido = "\n".join([";".join(p) for p in novas_linhas_parts])
    try:
        df_corr = pd.read_csv(StringIO(csv_corrigido), sep=";", dtype=str, engine="python", on_bad_lines="skip", quoting=0)
        return True, df_corr, csv_corrigido
    except Exception as e:
        print(f"[pagina_corrompida] Erro ao reler CSV corrigido: {e}"); return True, None, csv_corrigido

# ***** NOVA FUNÇÃO PARA LIMPEZA E TRANSFORMAÇÃO DOS DADOS *****
# NOVA VERSÃO CORRIGIDA DA FUNÇÃO
def processar_dataframe_final(df: pd.DataFrame, ordem_colunas: List[str]) -> pd.DataFrame:
    if df.empty:
        return df

    # 1. Remover linhas de rodapé (totalRegistros)
    primeira_coluna = df.columns[0]
    df = df[~df[primeira_coluna].astype(str).str.contains("totalRegistros|totalPaginas", case=False, na=False)]
    if df.empty:
        return df

    # 3. Criar a coluna "Unidade de Fornecimento" com a nova regra
    def criar_unidade_fornecimento(row):
        p1 = row.get('nomeUnidadeFornecimento')
        p2 = row.get('capacidadeUnidadeFornecimento')
        p3 = row.get('siglaUnidadeMedida')
        
        # Nova regra: só preenche se TODAS as 3 colunas tiverem conteúdo
        partes_validas = [
            str(p) for p in [p1, p2, p3] if pd.notna(p) and str(p).strip()
        ]
        
        if len(partes_validas) == 3:
            return " ".join(partes_validas)
        else:
            return "" # Retorna vazio se alguma parte faltar
    
    df['Unidade de Fornecimento'] = df.apply(criar_unidade_fornecimento, axis=1)

    # 4. Criar a coluna "Preço Total" com a lógica de conversão corrigida
    def converter_para_float(valor):
        if pd.isna(valor): return 0.0
        
        valor_str = str(valor)
        try:
            # Lógica corrigida: remove todos os pontos, depois troca a vírgula
            return float(valor_str.replace('.', '').replace(',', '.'))
        except (ValueError, TypeError):
            return 0.0

    preco_num = df['precoUnitario'].apply(converter_para_float)
    quantidade_num = df['quantidade'].apply(converter_para_float)
    df['Preço Total'] = preco_num * quantidade_num

    # 6. Excluir colunas se estiverem totalmente vazias
    for col in ["nomeUnidadeMedida", "percentualMaiorDesconto"]:
        if col in df.columns:
            if df[col].isnull().all() or df[col].astype(str).str.strip().eq('').all():
                df = df.drop(columns=[col])

    # 5. Reordenar colunas para a ordem final
    colunas_existentes_na_ordem = [col for col in ordem_colunas if col in df.columns]
    colunas_extras = [col for col in df.columns if col not in colunas_existentes_na_ordem]
    df = df[colunas_existentes_na_ordem + colunas_extras]

    return df

# --- ESTRUTURA DA INTERFACE GRÁFICA COM ABAS ---
sg.theme('SystemDefaultForReal')

welcome_message = """Olá! Bem-vindo ao Extrator de CATMATs Pro.

Sua ferramenta para extrair e descobrir dados no Portal de Compras Governamentais!

O que este programa faz?
Este programa possui duas funções principais em abas separadas:

1.  Extração por CATMAT (Aba 1): Se você já tem uma lista de códigos de materiais (CATMATs), esta aba busca todas as informações de compras, corrige problemas nos dados e consolida tudo em um arquivo Excel.

2.  Explorador de Classes (Aba 2): Se você quer descobrir novos itens, pode começar com o código de uma Classe, encontrar todos os Padrões Descritivos de Materiais (PDMs) dentro dela e, em seguida, listar todos os CATMATs relacionados para extração.

Primeiros Passos:
- Para uma extração direta com uma lista pronta, use a primeira aba.
- Para descobrir itens, comece pela segunda aba e, ao final, envie os CATMATs encontrados para a extração na primeira aba.

Acompanhe todo o processo em tempo real aqui neste log. Bom trabalho!
"""

# Layout da Aba 1: Extração Padrão
layout_extracao_config = [
    [sg.Text("Arquivo de Códigos:", size=(20,1)), sg.Input(key="-ARQUIVO-", enable_events=True, expand_x=True), sg.FileBrowse(button_text='Procurar', file_types=(("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")))],
    [sg.Checkbox('Salvar cópias dos arquivos CSV corrompidos?', key='-SALVAR_CORROMPIDOS-', default=False, enable_events=True)],
    [sg.Column([[sg.Text("Pasta para Corrompidos:", size=(20,1)), sg.Input(key="-PASTA-", enable_events=True, expand_x=True), sg.FolderBrowse(button_text='Procurar')]], key='-SECAO_PASTA_CORROMPIDOS-', visible=False)],
]
layout_extracao_stats = [
    [sg.Text("Códigos Processados:", size=(20,1)), sg.Text("0 / 0", key="-CONT_PROCESSADOS-", font=("Helvetica", 10, "bold"))],
    [sg.Text("Registros Consolidados:", size=(20,1)), sg.Text("0", key="-CONT_REGISTROS-", font=("Helvetica", 10, "bold"))],
    [sg.Text("Páginas Corrigidas:", size=(20,1)), sg.Text("0", key="-CONT_CORRIGIDAS-", font=("Helvetica", 10, "bold"), text_color="orange")],
    [sg.Text("Códigos sem Dados:", size=(20,1)), sg.Text("0", key="-CONT_VAZIOS-", font=("Helvetica", 10, "bold"), text_color="#FF6347")],
]
layout_extracao_execucao = [
    [sg.Text("Status: Ocioso", key='-STATUS-', expand_x=True, font=("Helvetica", 10, "italic"))],
    [sg.ProgressBar(max_value=1000, orientation='h', size=(50, 20), key='-PROGRESS-', expand_x=True), sg.Text('0%', size=(5,1), key='-PERCENT-', font=("Helvetica", 10, "bold"))],
    [sg.Multiline(default_text=welcome_message, size=(80,20), key="-OUTPUT-", autoscroll=True, expand_x=True, expand_y=True, background_color='black', text_color='white')],
]
tab1_layout = [
    [sg.Frame('1. Configurações de Entrada', layout_extracao_config, expand_x=True)],
    [sg.Frame('2. Resumo da Execução', layout_extracao_stats, expand_x=True)],
    [sg.Frame('3. Log e Progresso', layout_extracao_execucao, expand_x=True, expand_y=True)],
    [sg.Button("Iniciar Extração", key="-START-", disabled=True), sg.Button("Cancelar", key="-CANCEL-", disabled=True), sg.Button("Salvar Log", key="-SAVE_LOG-", disabled=True)]
]

# Cores para os botões de filtro
COR_BOTAO_SELECIONADO = ('white', 'green')
COR_BOTAO_PADRAO = ('white', 'grey')
FILTRO_BOTOES = ['-FILTRO_TODOS-', '-FILTRO_ATIVOS-', '-FILTRO_INATIVOS-']

# Layout da Aba 2: Explorador de Classes
pdm_headings = ['Código PDM', 'Descrição', 'Status']
layout_explorador = [
    [sg.Frame('1. Buscar PDMs por Classe', [[
        sg.Text('Código da Classe:'), 
        sg.Input(key='-INPUT_CLASSE-', size=(10,1)), 
        sg.Button('Buscar PDMs', key='-BUSCAR_PDMS-'),
        sg.Push(),
        sg.Text("", key="-PDM_COUNT_DISPLAY-", font=("Helvetica", 10, "bold"))
    ]], expand_x=True)],
    [sg.Frame('2. Resultados da Busca de PDMs', [[
        sg.Text("Filtro:"), 
        sg.Button('Todos', key='-FILTRO_TODOS-', button_color=COR_BOTAO_SELECIONADO), 
        sg.Button('Apenas Ativos', key='-FILTRO_ATIVOS-', button_color=COR_BOTAO_PADRAO), 
        sg.Button('Apenas Inativos', key='-FILTRO_INATIVOS-', button_color=COR_BOTAO_PADRAO),
        sg.Push(),
        sg.Checkbox('Selecionar Todos Visíveis', key='-SELECIONAR_TODOS_PDM-', enable_events=True)
        ], [
        sg.Table(values=[], headings=pdm_headings, num_rows=15, key='-TABELA_PDMS-', enable_events=True, justification='left', auto_size_columns=False, col_widths=[10, 50, 8], expand_x=True)
    ]], expand_x=True, expand_y=True)],
    [sg.Frame('3. Ações', [[
        sg.Button('Buscar CATMATs', key='-BUSCAR_CATMATS-'),
        sg.Button('Exportar Lista de CATMATs', key='-EXPORTAR_CATMATS-', disabled=True),
        sg.Column([[
            sg.Text("", key="-STATUS_EXPLORADOR-", font=("Helvetica", 10, "italic")),
            sg.Button('Cancelar Busca', key='-CANCELAR_BUSCA_CATMAT-', button_color=('white', 'red'))
        ]], key='-COLUNA_CANCELAR_BUSCA-', visible=False)
        ],[
        sg.Column([[
            sg.Button('Iniciar Extração com CATMATs Encontrados', key='-INICIAR_EXTRACAO_EXPLORADOR-', disabled=True, button_color=('white', 'green'), tooltip='Clique aqui para enviar a lista de CATMATs encontrados para a aba de extração e iniciar o processo.')
        ]], key='-COLUNA_INICIAR_EXTRACAO-', visible=False)
    ]], expand_x=True)],
]

# Layout principal com as abas
layout = [[sg.TabGroup([[
    sg.Tab('Extração por CATMAT', tab1_layout, key='-TAB_EXTRACAO-'),
    sg.Tab('Explorador de Classes', layout_explorador, key='-TAB_EXPLORADOR-')
]], key='-TAB_GROUP-', expand_x=True, expand_y=True)]]

window = sg.Window("Extrator de CATMATs Pro", layout, resizable=True, finalize=True)
window.set_min_size((800, 700))

window['-INPUT_CLASSE-'].bind('<Return>', '_Enter')

# --- LOOP DE EVENTOS E LÓGICA DE CONTROLE ---
URL_BASE = "https://dadosabertos.compras.gov.br"
TIMEOUT = 60
ordem_final_colunas = [
    "idCompra", "idItemCompra", "forma", "modalidade", "criterioJulgamento",
    "numeroItemCompra", "descricaoItem", "codigoItemCatalogo", "nomeUnidadeFornecimento",
    "siglaUnidadeFornecimento", "nomeUnidadeMedida",  "capacidadeUnidadeFornecimento", "siglaUnidadeMedida",
    "Unidade de Fornecimento", "capacidade", "quantidade", "precoUnitario", "Preço Total", "percentualMaiorDesconto",
    "niFornecedor", "nomeFornecedor", "marca", "codigoUasg", "nomeUasg",
    "codigoMunicipio", "municipio", "estado", "codigoOrgao", "nomeOrgao",
    "poder", "esfera", "dataCompra", "dataHoraAtualizacaoCompra", "dataHoraAtualizacaoItem",
    "dataResultado", "dataHoraAtualizacaoUasg", "codigoClasse", "nomeClasse"
]

# Variáveis de estado da extração
processing = False; codes_iterator = None; writer = None
codigos_para_processar, paginas_corrompidas, registros_esperados, registros_baixados = [], {}, {}, {}
total_registros_baixados, paginas_corrigidas_count, codigos_vazios_count = 0, 0, 0

# Variáveis de estado do explorador
lista_pdms_completa = pd.DataFrame()
lista_catmats_descobertos = []
cancelar_busca_catmat = False

# Função para a thread de busca de CATMATs
def buscar_catmats_thread(pdms_selecionados, window):
    df_catmats = buscar_catmats_por_pdm(pdms_selecionados, URL_BASE, TIMEOUT, window)
    window.write_event_value('-THREAD_CATMAT_CONCLUIDA-', df_catmats)

# Função para atualizar cores dos botões
def atualizar_cores_filtro(botao_ativo: str, window: sg.Window):
    for botao in FILTRO_BOTOES:
        if botao == botao_ativo:
            window[botao].update(button_color=COR_BOTAO_SELECIONADO)
        else:
            window[botao].update(button_color=COR_BOTAO_PADRAO)

# Função auxiliar para iniciar o processo de extração
def iniciar_processo_extracao(lista_codigos):
    global processing, codes_iterator, writer, paginas_corrompidas, registros_esperados, registros_baixados, codigos_para_processar
    global total_registros_baixados, paginas_corrigidas_count, codigos_vazios_count
    
    if not lista_codigos:
        sg.popup_error("Nenhum código para processar.")
        return

    processing = True
    codigos_para_processar = lista_codigos
    window["-START-"].update(disabled=True); window["-CANCEL-"].update(disabled=False); window["-SAVE_LOG-"].update(disabled=True)
    window["-OUTPUT-"].update(""); window["-PROGRESS-"].update(0); window["-PERCENT-"].update('0%')
    window["-CONT_PROCESSADOS-"].update(f"0 / {len(codigos_para_processar)}"); window["-CONT_REGISTROS-"].update("0")
    window["-CONT_CORRIGIDAS-"].update("0"); window["-CONT_VAZIOS-"].update("0")

    writer = ExcelChunkWriter("dados_completos_extraidos.xlsx")
    paginas_corrompidas, registros_esperados, registros_baixados = {}, {}, {}
    total_registros_baixados, paginas_corrigidas_count, codigos_vazios_count = 0, 0, 0
    
    codes_iterator = iter(enumerate(codigos_para_processar, 1))
    window["-STATUS-"].update("Status: Processando...")

while True:
    event, values = window.read(timeout=100)

    if event == sg.WIN_CLOSED: break

    # --- LÓGICA GERAL E DA ABA DE EXTRAÇÃO ---
    if not processing:
        arquivo_ok = bool(values['-ARQUIVO-'])
        pasta_ok = not values['-SALVAR_CORROMPIDOS-'] or (values['-SALVAR_CORROMPIDOS-'] and values['-PASTA-'])
        if arquivo_ok and pasta_ok: window['-START-'].update(disabled=False)
        else: window['-START-'].update(disabled=True)

    if event == '-SALVAR_CORROMPIDOS-':
        window['-SECAO_PASTA_CORROMPIDOS-'].update(visible=values['-SALVAR_CORROMPIDOS-'])
    
    if event == "-START-" and not processing:
        try:
            ARQUIVO_CODIGOS = values["-ARQUIVO-"]
            df_codigos = pd.read_excel(ARQUIVO_CODIGOS) if ARQUIVO_CODIGOS.lower().endswith(".xlsx") else pd.read_csv(ARQUIVO_CODIGOS, sep=";")
            if "codigoItemCatalogo" not in df_codigos.columns:
                sg.popup_error("Erro: O arquivo de entrada deve ter a coluna 'codigoItemCatalogo'.")
                continue
            lista_codigos = pd.Series(df_codigos["codigoItemCatalogo"]).dropna().astype(int).drop_duplicates().tolist()
            window["-OUTPUT-"].print(f"🔎 {len(lista_codigos)} códigos únicos carregados do arquivo.", text_color='lightblue')
            iniciar_processo_extracao(lista_codigos)
        except Exception as e:
            sg.popup_error(f"Ocorreu um erro ao ler o arquivo de códigos:\n{e}")

    if event == '-INICIAR_EXTRACAO_EXPLORADOR-':
        if not lista_catmats_descobertos:
            sg.popup_error("Nenhum CATMAT foi encontrado ou selecionado.")
        else:
            window['-TAB_EXTRACAO-'].select()
            window["-OUTPUT-"].print(f"🔎 {len(lista_catmats_descobertos)} códigos descobertos via explorador.", text_color='lightblue')
            iniciar_processo_extracao(lista_catmats_descobertos)

    if event == "-CANCEL-" and processing:
        processing = False; codes_iterator = None
        window["-STATUS-"].update("Status: Cancelando..."); window["-OUTPUT-"].print("\n🛑 Processo cancelado. Finalizando arquivos...", text_color='yellow')
        if writer:
            saved_parts = writer.finalize()
            if saved_parts: window["-OUTPUT-"].print(f"💾 Dados parciais salvos em: {', '.join(saved_parts)}", text_color='yellow')
        window["-PROGRESS-"].update(0); window["-PERCENT-"].update('0%');
        window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True)
        window["-SAVE_LOG-"].update(disabled=False); window["-STATUS-"].update("Status: Cancelado")

    if event == "-SAVE_LOG-":
        filepath = sg.popup_get_file("Salvar Log de Execução", save_as=True, no_window=True, default_extension=".txt", file_types=(("Text Files", "*.txt"),))
        if filepath:
            try:
                with open(filepath, 'w', encoding='utf-8') as f: f.write(values["-OUTPUT-"])
                sg.popup_quick("Log salvo com sucesso!")
            except Exception as e: sg.popup_error(f"Não foi possível salvar o log.\nErro: {e}")

    # --- LÓGICA DO EXPLORADOR DE CLASSES ---
    if event in ('-BUSCAR_PDMS-', '-INPUT_CLASSE-_Enter'):
        codigo_classe = values['-INPUT_CLASSE-']
        if codigo_classe and codigo_classe.isdigit():
            window['-STATUS_EXPLORADOR-'].update("Buscando PDMs... Aguarde.")
            window['-PDM_COUNT_DISPLAY-'].update("")
            window.refresh()
            
            resultado_busca = buscar_pdms_por_classe(int(codigo_classe), URL_BASE, TIMEOUT)
            
            if resultado_busca is not None:
                df_pdms, total_api = resultado_busca
                
                df_processado = df_pdms.rename(columns={'codigoPdm': 'Código PDM', 'nomePdm': 'Descrição', 'statusPdm': 'Status'})
                df_processado['Status'] = df_processado['Status'].apply(lambda x: 'Ativo' if x else 'Inativo')
                lista_pdms_completa = df_processado
                
                window['-TABELA_PDMS-'].update(values=lista_pdms_completa[['Código PDM', 'Descrição', 'Status']].values.tolist())
                
                total_encontrados = len(lista_pdms_completa)
                window['-PDM_COUNT_DISPLAY-'].update(f"{total_encontrados} de {total_api} PDMs encontrados.")
                window['-STATUS_EXPLORADOR-'].update("")
                
                atualizar_cores_filtro('-FILTRO_TODOS-', window)
                window['-SELECIONAR_TODOS_PDM-'].update(value=False)
                window['-EXPORTAR_CATMATS-'].update(disabled=True)
                window['-COLUNA_INICIAR_EXTRACAO-'].update(visible=False)
                window['-INICIAR_EXTRACAO_EXPLORADOR-'].update('Iniciar Extração com CATMATs Encontrados')
            else:
                window['-STATUS_EXPLORADOR-'].update("Nenhum PDM encontrado ou erro na busca.")
                window['-PDM_COUNT_DISPLAY-'].update("")
                lista_pdms_completa = pd.DataFrame()
                window['-TABELA_PDMS-'].update(values=[])
        elif event == '-BUSCAR_PDMS-':
            sg.popup_error("Por favor, insira um código de Classe válido (apenas números).")

    if event in FILTRO_BOTOES:
        if not lista_pdms_completa.empty:
            atualizar_cores_filtro(event, window)
            df_filtrado = lista_pdms_completa
            if event == '-FILTRO_ATIVOS-': df_filtrado = lista_pdms_completa[lista_pdms_completa['Status'] == 'Ativo']
            elif event == '-FILTRO_INATIVOS-': df_filtrado = lista_pdms_completa[lista_pdms_completa['Status'] == 'Inativo']
            window['-TABELA_PDMS-'].update(values=df_filtrado[['Código PDM', 'Descrição', 'Status']].values.tolist())
            window['-SELECIONAR_TODOS_PDM-'].update(value=False)

    if event == '-SELECIONAR_TODOS_PDM-':
        if values['-SELECIONAR_TODOS_PDM-']:
            num_rows = len(window['-TABELA_PDMS-'].Values)
            window['-TABELA_PDMS-'].update(select_rows=list(range(num_rows)))
        else:
            window['-TABELA_PDMS-'].update(select_rows=[])

    if event == '-BUSCAR_CATMATS-':
        indices_selecionados = window['-TABELA_PDMS-'].SelectedRows
        if not indices_selecionados: sg.popup_error("Selecione pelo menos um PDM na tabela.")
        else:
            dados_tabela = window['-TABELA_PDMS-'].Values
            pdms_selecionados = [int(dados_tabela[i][0]) for i in indices_selecionados if i < len(dados_tabela)]
            
            if not pdms_selecionados:
                sg.popup_error("A seleção de PDMs é inválida. Tente selecionar novamente.")
                continue
            
            cancelar_busca_catmat = False
            window['-BUSCAR_CATMATS-'].update(disabled=True)
            window['-COLUNA_CANCELAR_BUSCA-'].update(visible=True)
            
            threading.Thread(target=buscar_catmats_thread, args=(pdms_selecionados, window), daemon=True).start()

    if event == '-CANCELAR_BUSCA_CATMAT-':
        cancelar_busca_catmat = True

    if event == '-THREAD_CATMAT_CONCLUIDA-':
        df_catmats = values[event]
        window['-COLUNA_CANCELAR_BUSCA-'].update(visible=False)
        window['-BUSCAR_CATMATS-'].update(disabled=False)

        if df_catmats is not None and 'codigoItem' in df_catmats.columns:
            lista_catmats_descobertos = df_catmats['codigoItem'].dropna().astype(int).tolist()
            num_catmats = len(lista_catmats_descobertos)
            window['-STATUS_EXPLORADOR-'].update(f"✅ {num_catmats} CATMATs encontrados!")
            window['-INICIAR_EXTRACAO_EXPLORADOR-'].update(f'Iniciar Extração com {num_catmats} CATMATs Encontrados')
            window['-EXPORTAR_CATMATS-'].update(disabled=False)
            window['-COLUNA_INICIAR_EXTRACAO-'].update(visible=True)
            window['-INICIAR_EXTRACAO_EXPLORADOR-'].update(disabled=False)
        else:
            lista_catmats_descobertos = []
            if not cancelar_busca_catmat:
                window['-STATUS_EXPLORADOR-'].update("Nenhum CATMAT encontrado.")
            window['-EXPORTAR_CATMATS-'].update(disabled=True)
            window['-COLUNA_INICIAR_EXTRACAO-'].update(visible=False)
            window['-INICIAR_EXTRACAO_EXPLORADOR-'].update(disabled=True)

    if event == '-EXPORTAR_CATMATS-':
        if not lista_catmats_descobertos:
            sg.popup_error("Não há CATMATs na lista para exportar. Realize uma busca primeiro.")
        else:
            filepath = sg.popup_get_file(
                "Salvar Lista de CATMATs",
                save_as=True,
                no_window=True,
                default_extension=".csv",
                default_path="CATMATs_descobertos.csv",
                file_types=(("CSV Files", "*.csv"),)
            )
            if filepath:
                try:
                    df_export = pd.DataFrame(lista_catmats_descobertos, columns=['codigoItemCatalogo'])
                    df_export.to_csv(filepath, index=False, sep=';')
                    sg.popup_ok(f"Lista com {len(df_export)} CATMATs salva com sucesso em:\n{filepath}")
                except Exception as e:
                    sg.popup_error(f"Não foi possível salvar a lista.\nErro: {e}")

    # --- LÓGICA PRINCIPAL DE PROCESSAMENTO (QUANDO ATIVA) ---
    if processing:
        try:
            idx, codigo = next(codes_iterator)
            total_codigos = len(codigos_para_processar)
            window["-STATUS-"].update(f"Status: Processando código {codigo} ({idx}/{total_codigos})")
            window["-CONT_PROCESSADOS-"].update(f"{idx} / {total_codigos}")
            
            pagina_atual, total_paginas, baixados_do_codigo = 1, None, 0
            
            try:
                dfp, csv_text = ler_pagina_catmat(codigo, pagina_atual, URL_BASE, 500, TIMEOUT)
                if dfp is None or dfp.empty or csv_text is None:
                    window["-OUTPUT-"].print(f"ℹ️ Código {codigo}: Nenhum registro encontrado.", text_color='lightblue'); codigos_vazios_count += 1; window["-CONT_VAZIOS-"].update(codigos_vazios_count)
                    raise StopIteration 
                
                m_reg = re.search(r"totalRegistros\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                registros_esperados[codigo] = int(m_reg.group(1)) if m_reg else 0
                if registros_esperados[codigo] == 0:
                     window["-OUTPUT-"].print(f"ℹ️ Código {codigo}: API informa 0 registros.", text_color='lightblue'); codigos_vazios_count += 1; window["-CONT_VAZIOS-"].update(codigos_vazios_count)
                     raise StopIteration

                while True: 
                    is_corrupt, df_corrigido, _ = pagina_corrompida(csv_text, 19, 34)
                    df_final_pagina = None
                    if is_corrupt:
                        paginas_corrompidas.setdefault(codigo, []).append(str(pagina_atual))
                        if values['-SALVAR_CORROMPIDOS-'] and values['-PASTA-']:
                            path = os.path.join(values['-PASTA-'], f"cod_{codigo}_pag_{pagina_atual}_corrompido.csv")
                            with open(path, 'w', encoding='utf-8-sig') as f: f.write(csv_text)
                        
                        if df_corrigido is not None and not df_corrigido.empty:
                            window["-OUTPUT-"].print(f"⚠️ Cód {codigo}, Pág {pagina_atual}: Corrigida.", text_color='orange'); df_final_pagina = df_corrigido
                            paginas_corrigidas_count += 1; window["-CONT_CORRIGIDAS-"].update(paginas_corrigidas_count)
                        else: window["-OUTPUT-"].print(f"❌ Cód {codigo}, Pág {pagina_atual}: Falha na correção.", text_color='#FF6347')
                    else:
                        window["-OUTPUT-"].print(f"✅ Cód {codigo}, Pág {pagina_atual}: Lida com sucesso.", text_color='lightgreen'); df_final_pagina = dfp

                    if df_final_pagina is not None and not df_final_pagina.empty:
                        df_final_pagina.loc[:, "codigoItemCatalogo"] = str(codigo)
                        
                        df_processado = processar_dataframe_final(df_final_pagina, ordem_final_colunas)
                        
                        registros_na_pagina = len(df_processado)
                        baixados_do_codigo += registros_na_pagina
                        total_registros_baixados += registros_na_pagina
                        window["-CONT_REGISTROS-"].update(f"{total_registros_baixados:,}".replace(",", "."))
                        writer.write_dataframe(df_processado)
                    
                    if total_paginas is None:
                        m = re.search(r"total\s*p[áa]ginas?\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                        total_paginas = int(m.group(1)) if m else 1
                    
                    pagina_atual += 1
                    time.sleep(0.5)
                    if pagina_atual > total_paginas: break
                    dfp, csv_text = ler_pagina_catmat(codigo, pagina_atual, URL_BASE, 500, TIMEOUT)
                    if dfp is None or csv_text is None: break

            except StopIteration: pass
            except Exception as e: window["-OUTPUT-"].print(f"❌ Erro crítico no código {codigo}: {e}", text_color='#FF6347')
            registros_baixados[codigo] = baixados_do_codigo
            
            percent_complete = int((idx / total_codigos) * 100)
            window['-PROGRESS-'].update(percent_complete * 10)
            window['-PERCENT-'].update(f'{percent_complete}%')

        except StopIteration: 
            window['-PROGRESS-'].update(1000); window['-PERCENT-'].update('100%')
            window["-STATUS-"].update("Status: Processamento concluído! Gerando relatório...")
            window.refresh()
            
            processing = False
            window["-OUTPUT-"].print("\n🎉 Download concluído! Gerando arquivo final...", text_color='lightblue')

            saved_parts = writer.finalize()
            if not saved_parts:
                sg.popup("Nenhum dado válido foi baixado. O relatório não será gerado.")
                window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True); window["-SAVE_LOG-"].update(disabled=False); window["-STATUS-"].update("Status: Ocioso")
                
            else: 
                window["-OUTPUT-"].print(f"💾 Arquivos de dados salvos em: {', '.join(saved_parts)}", text_color='lightblue')
                
                ultimo_arquivo = saved_parts[-1]
                wb = load_workbook(ultimo_arquivo)
                if "Relatório Integridade" in wb.sheetnames: wb.remove(wb["Relatório Integridade"])
                ws_rel = wb.create_sheet("Relatório Integridade")
                ws_rel.append(["codigoItemCatalogo", "registros_esperados_api", "registros_baixados_reais", "paginas_com_problemas", "status"])
                for c in codigos_para_processar:
                    baixados = int(registros_baixados.get(c, 0))
                    esperados = int(registros_esperados.get(c, 0))
                    paginas = paginas_corrompidas.get(c, [])
                    status = "OK"
                    if esperados > 0 and baixados < esperados: status = f"Inconsistência (Baixados: {baixados} / Esperados: {esperados})"
                    elif esperados == 0 and baixados > 0: status = f"OK (API reportou 0, mas {baixados} foram encontrados)"
                    ws_rel.append([c, esperados, baixados, ", ".join(map(str, paginas)), status])
                wb.save(ultimo_arquivo)
                
                window["-OUTPUT-"].print("📊 Relatório de integridade adicionado.", text_color='lightblue')
                
                resumo_final = f"""
    ✅ Processo Concluído!
    -------------------------------------------
      - Códigos Processados: {len(codigos_para_processar)}
      - Registros Consolidados: {f"{total_registros_baixados:,}".replace(",", ".")}
      - Páginas Corrigidas: {paginas_corrigidas_count}
      - Códigos sem Registros: {codigos_vazios_count}
    -------------------------------------------"""
                sg.popup('Resumo da Extração', resumo_final)
                
                caminho_destino = sg.popup_get_file("Escolha onde salvar o arquivo final", save_as=True, no_window=True, default_path=os.path.basename(ultimo_arquivo), file_types=(("Arquivos Excel", "*.xlsx"),))
                if caminho_destino:
                    if not caminho_destino.lower().endswith(".xlsx"): caminho_destino += ".xlsx"
                    shutil.copy(ultimo_arquivo, caminho_destino); sg.popup("✅ Sucesso!", f"Arquivo final salvo em:\n{caminho_destino}")
                else: sg.popup("⚠ Atenção", f"Nenhum local escolhido. O arquivo final permanece em:\n{ultimo_arquivo}")
                
                window["-STATUS-"].update("Status: Concluído!"); window["-START-"].update(disabled=False); window["-CANCEL-"].update(disabled=True); window["-SAVE_LOG-"].update(disabled=False)

window.close()