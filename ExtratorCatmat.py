"""
Extrator de CATMATs Pro  —  v2.3
Motor gráfico: CustomTkinter  (tema claro/escuro nativo, cantos arredondados)
Identidade visual: inspirada no BPS / DESID (Gov.br)
"""

import re
import requests
import pandas as pd
from io import StringIO
from typing import Tuple, List, Optional
import os
import time
from openpyxl import Workbook
import shutil
import json
import threading
import math
from concurrent.futures import ThreadPoolExecutor, as_completed
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import customtkinter as ctk
import tkinter.ttk as ttk

# =============================================================================
# PALETA  —  neutros Gov.br + acento azul Gov + verde BPS + amarelo BPS
# =============================================================================
C_BG         = "#F4F5F7"   # cinza-papel (fundo geral)
C_SURFACE    = "#FFFFFF"   # branco (cards / frames)
C_BORDER     = "#DDE1E9"   # borda sutil
C_TEXT       = "#1A1D23"   # quase-preto
C_TEXT_MED   = "#555B6E"   # texto secundário
C_TEXT_LIGHT = "#8A92A6"   # placeholder / hint
C_ACCENT     = "#1351B4"   # azul Gov.br (primário)
C_ACCENT_H   = "#0C3784"   # hover do azul
C_GREEN      = "#168821"   # verde BPS (sucesso)
C_GREEN_H    = "#0E5C17"   # hover verde
C_YELLOW     = "#FFCD07"   # amarelo BPS (destaque / faixa)
C_ORANGE     = "#E37222"   # aviso
C_RED        = "#C0392B"   # erro / cancelar
C_LOG_BG     = "#13141A"   # terminal escuro
C_LOG_FG     = "#E8EAF0"   # texto terminal

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# =============================================================================
# LÓGICA DE NEGÓCIO
# =============================================================================

pausar_extracao     = threading.Event()
pausar_busca_catmat = threading.Event()
cancelar_busca_catmat = False

requests.packages.urllib3.disable_warnings(
    requests.packages.urllib3.exceptions.InsecureRequestWarning
)

# Session compartilhada — reutiliza conexões TCP/TLS entre todas as requisições
# Evita o overhead de handshake (~200-400ms) a cada chamada
_http = requests.Session()
_http.verify = False
_http.headers.update({"Accept-Encoding": "gzip, deflate", "Connection": "keep-alive"})
_adapter = requests.adapters.HTTPAdapter(
    pool_connections=4, pool_maxsize=12, max_retries=0
)
_http.mount("https://", _adapter)
_http.mount("http://",  _adapter)

URL_BASE = "https://dadosabertos.compras.gov.br"
TIMEOUT  = 120

ordem_final_colunas = [
    "idCompra","idItemCompra","forma","modalidade","criterioJulgamento",
    "numeroItemCompra","descricaoItem","codigoItemCatalogo","nomeUnidadeFornecimento",
    "siglaUnidadeFornecimento","nomeUnidadeMedida","capacidadeUnidadeFornecimento","siglaUnidadeMedida",
    "Unidade de Fornecimento","capacidade","quantidade","precoUnitario","Preco Total","percentualMaiorDesconto",
    "niFornecedor","nomeFornecedor","marca","codigoUasg","nomeUasg",
    "codigoMunicipio","municipio","estado","codigoOrgao","nomeOrgao",
    "poder","esfera","dataCompra","dataHoraAtualizacaoCompra","dataHoraAtualizacaoItem",
    "dataResultado","dataHoraAtualizacaoUasg","codigoClasse","nomeClasse",
]


class ExcelChunkWriter:
    def __init__(self, base_filename, sheet_name="Dados CATMAT", max_rows_per_file=1_000_000):
        self.base_filename = base_filename
        self.sheet_name    = sheet_name
        self.max_rows      = max_rows_per_file
        self.part          = 1
        self.header: List[str] = []
        self.current_row_count = 0
        self.files_saved: List[str] = []
        self._new_workbook()

    def _filepath(self):
        base, ext = os.path.splitext(self.base_filename)
        if not ext or ext.lower() != ".xlsx": ext = ".xlsx"
        return f"{base}_part{self.part}{ext}"

    def _new_workbook(self):
        self.wb = Workbook(); self.ws = self.wb.active
        self.ws.title = self.sheet_name
        self.header_written = False; self.current_row_count = 0

    def _ensure_header(self, columns):
        if not self.header: self.header = list(columns)
        if not self.header_written:
            self.ws.append(self.header); self.header_written = True

    def _rollover_if_needed(self):
        if self.current_row_count + 1 > self.max_rows:
            path = self._filepath(); self.wb.save(path); self.files_saved.append(path)
            self.part += 1; self._new_workbook()
            if self.header: self.ws.append(self.header); self.header_written = True

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        self._ensure_header(list(df.columns))
        for col in self.header:
            if col not in df.columns: df[col] = pd.NA
        df = df[self.header]
        for _, row in df.iterrows():
            self._rollover_if_needed()
            self.ws.append([None if pd.isna(v) else v for v in row])
            self.current_row_count += 1

    def finalize(self) -> List[str]:
        if self.header_written and self.current_row_count > 0:
            path = self._filepath(); self.wb.save(path)
            if path not in self.files_saved: self.files_saved.append(path)
        return self.files_saved


class CSVChunkWriter:
    def __init__(self, base_filename, sep=";", encoding="utf-8-sig", max_rows_per_file=1_000_000):
        self.base_filename = base_filename; self.sep = sep
        self.encoding = encoding; self.max_rows = max_rows_per_file
        self.part = 1; self.current_row_count = 0
        self.files_saved: List[str] = []; self.header_written = False

    def _filepath(self):
        base, ext = os.path.splitext(self.base_filename)
        if not ext or ext.lower() != ".csv": ext = ".csv"
        return f"{base}_part{self.part}{ext}"

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        if self.current_row_count + len(df) > self.max_rows:
            self.part += 1; self.current_row_count = 0; self.header_written = False
        path = self._filepath()
        df.to_csv(path, sep=self.sep, index=False,
                  mode="a" if self.header_written else "w",
                  header=not self.header_written, encoding=self.encoding)
        self.header_written = True; self.current_row_count += len(df)
        if path not in self.files_saved: self.files_saved.append(path)

    def finalize(self) -> List[str]:
        return self.files_saved


def converter_data_para_api(data_dd_mm_yyyy: str) -> Optional[str]:
    s = data_dd_mm_yyyy.strip()
    if not s: return None
    try:
        p = s.split("-")
        if len(p) != 3: return None
        dd, mm, yyyy = p
        if len(dd) == 2 and len(mm) == 2 and len(yyyy) == 4:
            int(dd); int(mm); int(yyyy)
            return f"{yyyy}-{mm}-{dd}"
    except (ValueError, AttributeError):
        pass
    return None


def validar_e_obter_datas(ini: str, fim: str):
    i_api = f_api = None
    if ini.strip():
        i_api = converter_data_para_api(ini)
        if i_api is None:
            return None, None, f"Data de Inicio invalida: '{ini}'\nUse DD-MM-AAAA (ex: 01-01-2024)"
    if fim.strip():
        f_api = converter_data_para_api(fim)
        if f_api is None:
            return None, None, f"Data Final invalida: '{fim}'\nUse DD-MM-AAAA (ex: 31-12-2024)"
    return i_api, f_api, None


def parse_csv_text(csv_text: str) -> pd.DataFrame:
    lines = [ln for ln in csv_text.splitlines() if ln.strip()]
    if not lines: return pd.DataFrame()
    try:
        return pd.read_csv(StringIO("\n".join(lines)), sep=";", dtype=str,
                           engine="python", on_bad_lines="warn", quoting=3)
    except Exception:
        return pd.DataFrame()


def ler_pagina_catmat(codigo, pagina, URL_BASE, TAMANHO_PAGINA, TIMEOUT,
                      data_compra_inicio=None, data_compra_fim=None):
    URL = f"{URL_BASE}/modulo-pesquisa-preco/1.1_consultarMaterial_CSV"
    params = {"tamanhoPagina": TAMANHO_PAGINA, "codigoItemCatalogo": int(codigo), "pagina": int(pagina)}
    if data_compra_inicio: params["dataCompraInicio"] = data_compra_inicio
    if data_compra_fim:    params["dataCompraFim"]    = data_compra_fim
    tentativas = 0
    while tentativas < 2:
        try:
            resp = _http.get(URL, params=params, timeout=TIMEOUT)
            if resp.status_code == 429:
                time.sleep(15 if tentativas == 0 else 30); tentativas += 1; continue
            resp.raise_for_status()
            return None, resp.content.decode("utf-8-sig", errors="replace")
        except requests.exceptions.ConnectionError as e:
            return None, f"ERRO_CONEXAO: {e}"
        except requests.exceptions.RequestException as e:
            return None, f"ERRO_REQUISICAO: {e}"
    return None, f"ERRO_REQUISICAO: 429 persistente para CATMAT {codigo}"


def _normalizar_campo(item: dict, *candidatos, default=""):
    """Retorna o primeiro campo encontrado no dict entre os candidatos."""
    for c in candidatos:
        if c in item and item[c] is not None:
            return item[c]
    return default


def buscar_pdms_por_classe(codigo_classe: int, URL_BASE: str, TIMEOUT: int,
                           max_tentativas: int = 3):
    """Busca todos os PDMs de uma classe com retry automático e backoff."""
    URL = f"{URL_BASE}/modulo-material/3_consultarPdmMaterial"
    TAMANHO_PAGINA = 500
    all_pdms = []; pagina_atual = 1; total_paginas = 1; total_registros_api = 0

    while pagina_atual <= total_paginas:
        tentativa = 0
        sucesso   = False
        data      = None
        while tentativa < max_tentativas and not sucesso:
            try:
                resp = _http.get(URL, params={
                    "codigoClasse": codigo_classe, "pagina": pagina_atual,
                    "tamanhoPagina": TAMANHO_PAGINA, "bps": "false"
                }, timeout=TIMEOUT)
                # Rate-limit: espera antes de tentar de novo
                if resp.status_code == 429:
                    espera = 15 * (tentativa + 1)
                    print(f"Rate-limit classe {codigo_classe} pág {pagina_atual} "
                          f"— aguardando {espera}s (tentativa {tentativa+1})")
                    time.sleep(espera)
                    tentativa += 1
                    continue
                resp.raise_for_status()
                data    = resp.json()
                sucesso = True
            except requests.exceptions.ConnectionError as e:
                espera = 3 * (tentativa + 1)
                print(f"Erro conexão classe {codigo_classe} pág {pagina_atual}: {e} "
                      f"— aguardando {espera}s (tentativa {tentativa+1})")
                time.sleep(espera)
                tentativa += 1
            except Exception as e:
                espera = 2 * (tentativa + 1)
                print(f"Erro classe {codigo_classe} pág {pagina_atual}: {e} "
                      f"— aguardando {espera}s (tentativa {tentativa+1})")
                time.sleep(espera)
                tentativa += 1

        if not sucesso or data is None:
            print(f"Falha definitiva: classe {codigo_classe} pág {pagina_atual} "
                  f"após {max_tentativas} tentativas")
            return None

        if "resultado" in data:
            all_pdms.extend(data["resultado"])
        if pagina_atual == 1:
            total_registros_api = int(data.get("totalRegistros", 0))
            total_paginas = (math.ceil(total_registros_api / TAMANHO_PAGINA)
                             if total_registros_api > 0 else 1)
            print(f"Classe {codigo_classe}: {total_registros_api} PDMs / "
                  f"{total_paginas} página(s)")
        pagina_atual += 1
        time.sleep(0.2)

    if not all_pdms: return None

    # Normalizar campos — a API pode retornar nomes variados
    rows_norm = []
    for item in all_pdms:
        cod  = _normalizar_campo(item, "codigoPdm", "codigo", "id", "codigoItem")
        desc = _normalizar_campo(item, "nomePdm", "nome", "descricao", "descricaoPdm", "descricaoItem")
        # status pode ser bool True/False, string "ATIVO"/"INATIVO", ou inteiro
        raw_status = _normalizar_campo(item, "statusPdm", "status", "ativo", "situacao")
        if isinstance(raw_status, bool):
            status = "Ativo" if raw_status else "Inativo"
        elif isinstance(raw_status, str):
            status = "Ativo" if raw_status.upper() in ("ATIVO", "TRUE", "S", "SIM", "1") else "Inativo"
        elif isinstance(raw_status, (int, float)):
            status = "Ativo" if raw_status == 1 else "Inativo"
        else:
            status = "Ativo"
        rows_norm.append({"codigoPdm": cod, "nomePdm": desc, "statusPdm": status,
                          "_classe": str(codigo_classe)})

    df = pd.DataFrame(rows_norm).drop_duplicates(subset=["codigoPdm"])
    return df, total_registros_api


def buscar_catmats_por_pdm(codigos_pdm, URL_BASE, TIMEOUT, app,
                           log_fn=None, max_workers=5):
    """
    Busca CATMATs de múltiplos PDMs em paralelo usando ThreadPoolExecutor.
    max_workers=5 → ~5x mais rápido sem sobrecarregar a API.
    """
    global cancelar_busca_catmat
    URL   = f"{URL_BASE}/modulo-material/4_consultarItemMaterial"
    total = len(codigos_pdm)
    all_catmats  = []
    pdms_com_erro = []
    completed     = 0
    lock          = threading.Lock()   # protege all_catmats e pdms_com_erro

    def _fetch_pdm(idx_pdm):
        """Worker: busca todas as páginas de CATMATs de um PDM."""
        i, pdm_code = idx_pdm
        if cancelar_busca_catmat:
            return
        resultados = []; pagina_atual = 1; total_paginas = 1
        try:
            while pagina_atual <= total_paginas:
                if cancelar_busca_catmat: break
                resp = _http.get(URL, params={
                    "codigoPdm": pdm_code, "pagina": pagina_atual,
                    "tamanhoPagina": 500, "bps": "false"
                }, timeout=TIMEOUT)
                if resp.status_code == 429:
                    time.sleep(15); continue          # rate-limit: aguarda e repete
                resp.raise_for_status()
                data = resp.json()
                if "resultado" in data:
                    resultados.extend(data["resultado"])
                if pagina_atual == 1:
                    total_paginas = data.get("totalPaginas", 1)
                pagina_atual += 1
                if pagina_atual <= total_paginas:
                    time.sleep(0.2)                   # throttle entre páginas do mesmo PDM
        except Exception:
            with lock:
                pdms_com_erro.append(pdm_code)
            return
        if resultados:
            with lock:
                all_catmats.extend(resultados)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_fetch_pdm, (i, pdm)): (i, pdm)
            for i, pdm in enumerate(codigos_pdm)
        }
        for future in as_completed(futures):
            pausar_busca_catmat.wait()        # respeita pausa
            if cancelar_busca_catmat:
                app.after(0, lambda: app.set_status_explorador("Busca cancelada."))
                executor.shutdown(wait=False, cancel_futures=True)
                break
            completed += 1
            i, pdm_code = futures[future]
            msg = f"PDM {pdm_code} ({completed}/{total})..."
            app.after(0, lambda m=msg: app.set_status_explorador(m))
            if log_fn and (completed % 5 == 0 or completed == total):
                app.after(0, lambda m=msg: log_fn(m))

    return pd.DataFrame(all_catmats) if all_catmats else None, pdms_com_erro



def _fetch_catmat_registros(codigo, d_ini, d_fim, salvar_corr, pasta_corr):
    """
    Worker puro: busca e processa todas as páginas de um CATMAT.
    Pode rodar em qualquer thread — não acessa estado compartilhado.
    Retorna: (codigo, dfs_e_meta, tipo, reg_esp, pag_corr)
      tipo: "ok" | "vazio" | "erro" | "conexao"
      dfs_e_meta: lista de (df_processado, is_corrompida, num_pagina)
    """
    dfs_e_meta  = []
    pag_corr    = {}
    reg_esp     = 0
    pagina_atual = 1; total_paginas = None

    try:
        _, csv_text = ler_pagina_catmat(codigo, 1, URL_BASE, 500, TIMEOUT, d_ini, d_fim)
        if csv_text and csv_text.startswith("ERRO_CONEXAO"):
            return codigo, [], "conexao", 0, {}
        if csv_text is None or csv_text.startswith("ERRO_REQUISICAO"):
            return codigo, [], "erro", 0, {}

        m = re.search(r"totalRegistros\s*:\s*(\d+)", csv_text, re.IGNORECASE)
        reg_esp = int(m.group(1)) if m else 0
        if reg_esp == 0:
            return codigo, [], "vazio", 0, {}

        while True:
            is_c, csv_c = pagina_corrompida(csv_text)
            df_pag = parse_csv_text(csv_c)

            if is_c:
                pag_corr.setdefault(codigo, []).append(str(pagina_atual))
                if salvar_corr and pasta_corr:
                    dest = os.path.join(pasta_corr,
                        f"cod_{codigo}_pag_{pagina_atual}_corr.csv")
                    try:
                        with open(dest, "w", encoding="utf-8-sig") as f:
                            f.write(csv_text)
                    except Exception:
                        pass

            if df_pag is not None and not df_pag.empty:
                df_pag.loc[:, "codigoItemCatalogo"] = str(codigo)
                df_proc = processar_dataframe_final(df_pag, ordem_final_colunas)
                dfs_e_meta.append((df_proc, is_c, pagina_atual))

            if total_paginas is None:
                mp = re.search(r"total\s*p[áa]ginas?\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                total_paginas = int(mp.group(1)) if mp else 1

            pagina_atual += 1
            if pagina_atual > total_paginas:
                break
            time.sleep(0.5)
            _, csv_text = ler_pagina_catmat(codigo, pagina_atual, URL_BASE, 500, TIMEOUT,
                                            d_ini, d_fim)
            if csv_text is None or csv_text.startswith("ERRO_"):
                break

        return codigo, dfs_e_meta, ("ok" if dfs_e_meta else "vazio"), reg_esp, pag_corr

    except Exception:
        return codigo, [], "erro", 0, {}


def pagina_corrompida(csv_text: str):
    if not csv_text: return False, csv_text
    linhas = [ln for ln in csv_text.splitlines() if ln.strip()]
    if not linhas: return False, csv_text
    try:
        hi = next(i for i, ln in enumerate(linhas)
                  if not ln.lower().startswith(("totalregistros:", "totalpaginas:")))
        header_line = linhas[hi]; ncols = len(header_line.split(";"))
    except StopIteration:
        return False, csv_text
    if ncols == 0: return False, csv_text
    out = list(linhas[:hi]) + [header_line]
    buf = ""; corrigido = False
    for ln in linhas[hi+1:]:
        if ln.lower().startswith(("totalregistros:", "totalpaginas:")):
            if buf: out.append(buf); buf = ""
            out.append(ln); continue
        atual = buf + ln.replace("\r","").replace("\n","")
        if len(atual.split(";")) < ncols:
            buf = atual + " "; corrigido = True
        else:
            out.append(atual); buf = ""
    if buf: out.append(buf)
    return corrigido, "\n".join(out)


def processar_dataframe_final(df: pd.DataFrame, ordem_colunas: List[str]) -> pd.DataFrame:
    if df.empty: return df
    fc = df.columns[0]
    df = df[~df[fc].astype(str).str.contains("totalRegistros|totalPaginas",
                                              case=False, na=False)].copy()
    if df.empty: return df

    def uf(row):
        # Usa as 4 colunas na ordem definida; inclui só as que têm valor
        partes = []
        for k in ["nomeUnidadeFornecimento", "siglaUnidadeFornecimento",
                  "capacidadeUnidadeFornecimento", "siglaUnidadeMedida"]:
            v = row.get(k)
            if pd.notna(v) and str(v).strip() and str(v).strip() not in ("nan", "None"):
                partes.append(str(v).strip())
        return " ".join(partes)

    df["Unidade de Fornecimento"] = df.apply(uf, axis=1)

    def tof(v):
        if pd.isna(v): return 0.0
        try: return float(str(v).replace(".", "").replace(",", "."))
        except: return 0.0

    df["Preco Total"] = df["precoUnitario"].apply(tof) * df["quantidade"].apply(tof)
    for col in ["nomeUnidadeMedida","percentualMaiorDesconto"]:
        if col in df.columns and (df[col].isnull().all() or
                                   df[col].astype(str).str.strip().eq("").all()):
            df = df.drop(columns=[col])
    exist = [c for c in ordem_colunas if c in df.columns]
    extra = [c for c in df.columns if c not in exist]
    return df[exist + extra]


# =============================================================================
# COMPONENTES DE UI  (helpers)
# =============================================================================

WELCOME = """\
Olá! Bem-vindo ao Extrator de CATMATs Pro.

Sua ferramenta para extrair e descobrir dados no Portal de Compras Governamentais!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
O que este programa faz?
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Este programa possui duas funcoes principais em abas separadas:

  1. Extracao por CATMAT (esta aba)
     Se voce ja tem uma lista de codigos de materiais (CATMATs), esta aba
     busca todas as informacoes de compras, corrige problemas nos dados e
     consolida tudo em um arquivo Excel ou CSV.

  2. Extracao por Classes (aba ao lado)
     Se voce quer descobrir novos itens, pode comecar com o codigo de uma
     ou mais Classes, encontrar todos os Padroes Descritivos de Materiais
     (PDMs) dentro delas e, em seguida, listar todos os CATMATs relacionados
     para extracao.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Primeiros Passos
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  - Para uma extracao direta com uma lista pronta, use esta aba.
    O arquivo Excel ou CSV deve ter a coluna: codigoItemCatalogo

  - Para descobrir itens, use a aba "Extracao por Classes" e, ao final,
    envie os CATMATs encontrados para a extracao nesta aba.

  - Utilize os filtros de data (DD-MM-AAAA) para restringir os resultados
    a um periodo especifico de compras (Data de Inicio e Data Final).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Acompanhe todo o processo em tempo real neste log. Bom trabalho!
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""


def _lbl(parent, text, size=12, weight="normal", color=C_TEXT, **kw):
    return ctk.CTkLabel(parent, text=text, font=("Segoe UI", size, weight),
                        text_color=color, **kw)


def _btn(parent, text, command, variant="secondary", width=0, **kw):
    pal = {
        "primary":   (C_SURFACE,  C_ACCENT,  C_SURFACE,  C_ACCENT_H),
        "success":   (C_SURFACE,  C_GREEN,   C_SURFACE,  C_GREEN_H),
        "danger":    (C_SURFACE,  C_RED,     C_SURFACE,  "#992B1E"),
        "secondary": (C_TEXT,     "#E4E7EF", C_TEXT,     C_BORDER),
        "ghost":     (C_ACCENT,   "transparent", C_ACCENT_H, "#E8EDF8"),
    }
    tc, bg, htc, hbg = pal.get(variant, pal["secondary"])
    return ctk.CTkButton(parent, text=text, command=command,
                         font=("Segoe UI", 12), fg_color=bg, text_color=tc,
                         hover_color=hbg, corner_radius=6,
                         width=width, height=32, **kw)


def _entry(parent, textvariable=None, placeholder="", width=200, **kw):
    return ctk.CTkEntry(parent, textvariable=textvariable,
                        placeholder_text=placeholder,
                        font=("Segoe UI", 12),
                        fg_color=C_SURFACE, text_color=C_TEXT,
                        border_color=C_BORDER, border_width=1,
                        corner_radius=6, width=width,
                        placeholder_text_color=C_TEXT_LIGHT, **kw)


def _sep(parent, pady=(6,6)):
    ctk.CTkFrame(parent, height=1, fg_color=C_BORDER,
                 corner_radius=0).pack(fill="x", padx=14, pady=pady)


def _card(parent, title="", **kw):
    outer = ctk.CTkFrame(parent, fg_color=C_SURFACE, corner_radius=8,
                         border_width=1, border_color=C_BORDER, **kw)
    if title:
        _lbl(outer, title, size=11, weight="bold", color=C_TEXT_MED)\
            .pack(anchor="w", padx=14, pady=(10,4))
        _sep(outer, pady=(0,6))
    return outer


# =============================================================================
# APLICATIVO PRINCIPAL
# =============================================================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Extrator de CATMATs Pro  |  BPS / DESID")
        self.withdraw()                          # esconde até centralizar
        self.geometry("1100x860")
        self.minsize(960, 720)
        self.configure(fg_color=C_BG)

        # estado
        self.processing           = False
        self.codes_iterator       = None
        self.writer               = None
        self.codigos_lista: List  = []
        self.paginas_corrompidas  = {}
        self.registros_esperados  = {}
        self.registros_baixados   = {}
        self.total_baixados       = 0
        self.count_corrigidas     = 0
        self.count_vazios         = 0
        self._data_inicio         = None
        self._data_fim            = None
        self.lista_pdms_df        = pd.DataFrame()
        self.lista_catmats: List  = []

        # Construir interface ANTES de centralizar
        self._build_header()
        self._build_tabs()

        # Centralizar após tudo construído — delay generoso para o Tkinter
        # calcular dimensões reais antes de exibir
        self.after(200, self._centralizar)

    def _centralizar(self):
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        # winfo_width retorna 1 até a janela aparecer; usar reqwidth como fallback
        ww = self.winfo_reqwidth()  or 1100
        wh = self.winfo_reqheight() or 800
        # Respeitar o geometry definido (1100x800)
        ww = max(ww, 1100)
        wh = max(wh, 860)
        x = max(0, (sw - ww) // 2)
        y = max(0, (sh - wh) // 2)
        self.geometry(f"{ww}x{wh}+{x}+{y}")
        self.deiconify()

    # ── HEADER ────────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=C_ACCENT, corner_radius=0, height=50)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        _lbl(hdr, "  Extrator de CATMATs Pro", size=15, weight="bold",
             color=C_SURFACE).pack(side="left", padx=6)
        _lbl(hdr, "BPS · DESID · Ministério da Saúde  ",
             size=10, color="#A8BFDF").pack(side="right")
        # faixa amarela
        ctk.CTkFrame(self, height=3, fg_color=C_YELLOW,
                     corner_radius=0).pack(fill="x")

    # ── TABS ──────────────────────────────────────────────────────────────────
    def _build_tabs(self):
        self.tabs = ctk.CTkTabview(
            self, fg_color=C_BG,
            segmented_button_fg_color=C_BORDER,
            segmented_button_selected_color=C_ACCENT,
            segmented_button_selected_hover_color=C_ACCENT_H,
            segmented_button_unselected_color=C_BORDER,
            segmented_button_unselected_hover_color="#C5CAD5",
            text_color=C_TEXT, text_color_disabled=C_TEXT_LIGHT,
            corner_radius=0)
        self.tabs.pack(fill="both", expand=True)
        self.tabs.add("  Extração por CATMAT  ")
        self.tabs.add("  Extração por Classes  ")
        self._build_tab_extracao(self.tabs.tab("  Extração por CATMAT  "))
        self._build_tab_explorador(self.tabs.tab("  Extração por Classes  "))

    # ── ABA 1 ─────────────────────────────────────────────────────────────────
    def _build_tab_extracao(self, parent):
        parent.configure(fg_color=C_BG)
        # Frame normal sem scroll — tudo deve caber na tela
        wrap = ctk.CTkFrame(parent, fg_color=C_BG, corner_radius=0)
        wrap.pack(fill="both", expand=True, padx=12, pady=8)

        # — Card 1: entrada —
        c1 = _card(wrap, "1.  Dados para a Extração")
        c1.pack(fill="x", pady=(0,8))
        inn = ctk.CTkFrame(c1, fg_color="transparent")
        inn.pack(fill="x", padx=14, pady=(0,12))

        # arquivo
        r = ctk.CTkFrame(inn, fg_color="transparent"); r.pack(fill="x", pady=3)
        _lbl(r, "Arquivo de Códigos:", color=C_TEXT_MED).pack(side="left", padx=(0,8))
        self.var_arquivo = tk.StringVar()
        _entry(r, textvariable=self.var_arquivo,
               placeholder="Selecione .xlsx ou .csv", width=420)\
            .pack(side="left", expand=True, fill="x")
        _btn(r, "Procurar…", self._escolher_arquivo, variant="ghost", width=90)\
            .pack(side="left", padx=(8,0))

        _sep(inn)

        # datas
        r2 = ctk.CTkFrame(inn, fg_color="transparent"); r2.pack(fill="x", pady=3)
        _lbl(r2, "Data de Início:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_ini1 = tk.StringVar()
        _entry(r2, textvariable=self.var_ini1, placeholder="DD-MM-AAAA", width=130)\
            .pack(side="left")
        _lbl(r2, "Data Final:", color=C_TEXT_MED).pack(side="left", padx=(20,6))
        self.var_fim1 = tk.StringVar()
        _entry(r2, textvariable=self.var_fim1, placeholder="DD-MM-AAAA", width=130)\
            .pack(side="left")

        _sep(inn)

        # formato
        r3 = ctk.CTkFrame(inn, fg_color="transparent"); r3.pack(fill="x", pady=3)
        _lbl(r3, "Formato de Saída:", color=C_TEXT_MED).pack(side="left", padx=(0,12))
        self.var_fmt = tk.StringVar(value="xlsx")
        for txt, val in [("Excel (.xlsx)","xlsx"), ("CSV (.csv)","csv")]:
            ctk.CTkRadioButton(r3, text=txt, variable=self.var_fmt, value=val,
                               font=("Segoe UI",12), text_color=C_TEXT,
                               fg_color=C_ACCENT, border_color=C_BORDER)\
                .pack(side="left", padx=(0,16))

        _sep(inn)

        # corrompidos
        r4 = ctk.CTkFrame(inn, fg_color="transparent"); r4.pack(fill="x", pady=3)
        self.var_salvar_corr = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(r4, text="Salvar cópias dos CSV corrompidos",
                        variable=self.var_salvar_corr, command=self._toggle_pasta,
                        font=("Segoe UI",12), text_color=C_TEXT,
                        fg_color=C_ACCENT, border_color=C_BORDER)\
            .pack(side="left")
        self.frame_pasta = ctk.CTkFrame(inn, fg_color="transparent")
        rp = ctk.CTkFrame(self.frame_pasta, fg_color="transparent")
        rp.pack(fill="x", pady=3)
        _lbl(rp, "Pasta:", color=C_TEXT_MED).pack(side="left", padx=(0,8))
        self.var_pasta = tk.StringVar()
        _entry(rp, textvariable=self.var_pasta,
               placeholder="Pasta de destino", width=380)\
            .pack(side="left", expand=True, fill="x")
        _btn(rp, "Procurar…", self._escolher_pasta, variant="ghost", width=90)\
            .pack(side="left", padx=(8,0))

        # — Card 2: estatísticas —
        c2 = _card(wrap, "2.  Resumo da Execução")
        c2.pack(fill="x", pady=(0,8))
        grid = ctk.CTkFrame(c2, fg_color="transparent")
        grid.pack(fill="x", padx=14, pady=(0,8))
        stats = [
            ("Códigos Processados",   "k_proc",  C_ACCENT),
            ("Registros Consolidados","k_reg",   C_GREEN),
            ("Páginas Corrigidas",    "k_corr",  C_ORANGE),
            ("Códigos sem Dados",     "k_vaz",   C_RED),
        ]
        self._stats = {}
        for col, (nome, key, cor) in enumerate(stats):
            cell = ctk.CTkFrame(grid, fg_color=C_BG, corner_radius=6,
                                border_width=1, border_color=C_BORDER)
            cell.grid(row=0, column=col, padx=5, pady=4, sticky="ew")
            grid.grid_columnconfigure(col, weight=1)
            _lbl(cell, nome, size=10, color=C_TEXT_MED).pack(pady=(6,1))
            lv = ctk.CTkLabel(cell, text="0", font=("Segoe UI",17,"bold"),
                              text_color=cor)
            lv.pack(pady=(0,6))
            self._stats[key] = lv

        # — Card 3: log —
        c3 = _card(wrap, "3.  Log e Progresso")
        c3.pack(fill="x", pady=(0,4))

        brow = ctk.CTkFrame(c3, fg_color="transparent")
        brow.pack(fill="x", padx=14, pady=(0,4))
        self.lbl_status = _lbl(brow, "Status: Ocioso", size=11,
                                color=C_TEXT_MED, anchor="w")
        self.lbl_status.pack(side="left", expand=True, fill="x")
        self.lbl_pct = _lbl(brow, "0%", size=11, weight="bold", color=C_GREEN)
        self.lbl_pct.pack(side="right", padx=(8,0))

        self.progress = ctk.CTkProgressBar(c3, fg_color=C_BORDER,
                                           progress_color=C_GREEN,
                                           corner_radius=3, height=6)
        self.progress.set(0)
        self.progress.pack(fill="x", padx=14, pady=(0,8))

        log_wrap = ctk.CTkFrame(c3, fg_color=C_LOG_BG, corner_radius=6)
        log_wrap.pack(fill="x", padx=14, pady=(0,10))
        self.log = scrolledtext.ScrolledText(
            log_wrap, bg=C_LOG_BG, fg=C_LOG_FG,
            font=("Consolas",10), wrap="word",
            relief="flat", bd=0, state="normal",
            height=9,
            insertbackground=C_LOG_FG)
        self.log.pack(fill="x", padx=6, pady=6)
        for tag, cor in [("ok","#4EC94E"),("warn","#F4A11D"),
                         ("err","#E05C5C"),("info","#7EB8F7"),
                         ("date","#FFCD07")]:
            self.log.tag_config(tag, foreground=cor)
        self._log(WELCOME, "info")

        # botões ficam no wrap (fora do card), sempre visíveis
        br = ctk.CTkFrame(wrap, fg_color="transparent")
        br.pack(fill="x", pady=(4,4))
        self.btn_start = _btn(br, "▶  Iniciar Extração", self._start,
                              variant="primary", width=160)
        self.btn_start.pack(side="left", padx=(0,8))
        self.btn_cancel = _btn(br, "✖  Cancelar", self._cancelar,
                               variant="secondary", width=100)
        self.btn_cancel.configure(state="disabled")
        self.btn_cancel.pack(side="left", padx=(0,8))
        self.btn_pause = _btn(br, "⏸  Pausar", self._pausar,
                              variant="secondary", width=100)
        self.btn_pause.configure(state="disabled")
        self.btn_pause.pack(side="left", padx=(0,8))
        self.btn_log = _btn(br, "💾  Salvar Log", self._salvar_log,
                            variant="secondary", width=120)
        self.btn_log.configure(state="disabled")
        self.btn_log.pack(side="left")

    # ── ABA 2 ─────────────────────────────────────────────────────────────────
    def _build_tab_explorador(self, parent):
        parent.configure(fg_color=C_BG)

        # ── Card 1: Classes (topo) ────────────────────────────────────────────
        c1 = _card(parent, "1.  Buscar PDMs por Classes")
        c1.pack(fill="x", padx=12, pady=(8,6))
        inn1 = ctk.CTkFrame(c1, fg_color="transparent")
        inn1.pack(fill="x", padx=14, pady=(0,10))
        _lbl(inn1, "Informe as Classes para extração separadas por  ;",
             color=C_TEXT_MED, size=11).pack(anchor="w", pady=(0,4))
        r = ctk.CTkFrame(inn1, fg_color="transparent")
        r.pack(fill="x")
        self.var_classe = tk.StringVar()
        ent = _entry(r, textvariable=self.var_classe,
                     placeholder="ex.: 20115 ; 20116 ; 20117", width=400)
        ent.pack(side="left", expand=True, fill="x")
        ent.bind("<Return>", lambda e: self._buscar_pdms())
        _btn(r, "Buscar PDMs", self._buscar_pdms,
             variant="primary", width=140).pack(side="left", padx=(10,0))
        _btn(r, "⚡  Buscar e Extrair", self._buscar_e_extrair_classes,
             variant="success", width=160).pack(side="left", padx=(8,0))
        self.lbl_pdm_count = _lbl(r, "", size=11, color=C_GREEN)
        self.lbl_pdm_count.pack(side="right", padx=8)
        row_chk = ctk.CTkFrame(inn1, fg_color="transparent")
        row_chk.pack(fill="x", pady=(8,0))
        self.var_arquivo_por_classe = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(row_chk,
                        text="Deseja salvar um arquivo por classe?",
                        variable=self.var_arquivo_por_classe,
                        command=self._toggle_pasta_por_classe,
                        font=("Segoe UI", 12), text_color=C_TEXT,
                        fg_color=C_ACCENT, border_color=C_BORDER)            .pack(side="left")
        _lbl(row_chk, "  (gera um arquivo separado para cada classe informada)",
             size=10, color=C_TEXT_LIGHT).pack(side="left")
        # Linha de pasta de destino — visível só quando checkbox marcado
        self.frame_pasta_classes = ctk.CTkFrame(inn1, fg_color="transparent")
        row_pasta = ctk.CTkFrame(self.frame_pasta_classes, fg_color="transparent")
        row_pasta.pack(fill="x")
        _lbl(row_pasta, "Pasta de destino:", color=C_TEXT_MED, size=11)            .pack(side="left", padx=(0,8))
        self.var_pasta_classes = tk.StringVar()
        _entry(row_pasta, textvariable=self.var_pasta_classes,
               placeholder="Selecione a pasta onde os arquivos serão salvos",
               width=380).pack(side="left", expand=True, fill="x")
        _btn(row_pasta, "📂  Procurar", self._escolher_pasta_classes,
             variant="ghost", width=100).pack(side="left", padx=(8,0))

        # ── Área central: tabela (esquerda) + busca avulsa (direita) ─────────
        mid = ctk.CTkFrame(parent, fg_color=C_BG)
        mid.pack(fill="both", expand=True, padx=12, pady=(0,4))

        # Painel lateral DIREITO: Busca Avulsa por PDMs
        # — deve ser empacotado ANTES do c2 para reservar espaço antes do expand
        cav = _card(mid, "Busca Avulsa por PDMs")
        cav.pack(side="right", fill="y", padx=(6,0))
        _lbl(cav, "Códigos PDM (um por linha):",
             size=11, color=C_TEXT_MED).pack(anchor="w", padx=14, pady=(0,4))
        self.txt_avulso = ctk.CTkTextbox(cav, font=("Consolas",11),
                                         fg_color=C_SURFACE, text_color=C_TEXT,
                                         border_width=1, border_color=C_BORDER,
                                         width=190, corner_radius=6)
        self.txt_avulso.pack(fill="both", expand=True, padx=14)
        _btn(cav, "🔍  Buscar CATMATs\n(PDMs da lista)",
             self._buscar_avulso, variant="primary")            .pack(fill="x", padx=14, pady=(8,10))

        # Card 2: PDMs Encontrados — expande para ocupar o espaço restante
        c2 = _card(mid, "2.  PDMs Encontrados")
        c2.pack(side="left", fill="both", expand=True)

        frow = ctk.CTkFrame(c2, fg_color="transparent")
        frow.pack(fill="x", padx=14, pady=(0,4))
        _lbl(frow, "Filtro:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_filtro = tk.StringVar(value="todos")
        for txt, val in [("Todos","todos"),("Ativos","ativo"),("Inativos","inativo")]:
            ctk.CTkRadioButton(frow, text=txt, variable=self.var_filtro, value=val,
                               command=self._filtrar,
                               font=("Segoe UI",11), text_color=C_TEXT,
                               fg_color=C_ACCENT, border_color=C_BORDER)                .pack(side="left", padx=(0,10))
        _btn(frow, "🔍  Buscar CATMATs (PDMs da tabela)", self._buscar_catmats,
             variant="secondary", width=240).pack(side="left", padx=(16,0))
        _btn(frow, "Exportar PDMs", self._exp_pdms,
             variant="ghost", width=120).pack(side="right")

        # Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("BPS.Treeview", background=C_SURFACE, foreground=C_TEXT,
                        fieldbackground=C_SURFACE, rowheight=26,
                        font=("Segoe UI",10), borderwidth=0)
        style.configure("BPS.Treeview.Heading", background=C_BG,
                        foreground=C_TEXT_MED, font=("Segoe UI",10,"bold"),
                        relief="flat")
        style.map("BPS.Treeview",
                  background=[("selected", C_ACCENT)],
                  foreground=[("selected", C_SURFACE)])

        tf = ctk.CTkFrame(c2, fg_color=C_SURFACE, corner_radius=0)
        tf.pack(fill="both", expand=True, padx=8, pady=(0,8))
        vsb = ttk.Scrollbar(tf, orient="vertical", command=None)
        vsb.pack(side="right", fill="y")
        self.tree = ttk.Treeview(tf, columns=("cod","desc","status"),
                                 show="headings", style="BPS.Treeview",
                                 selectmode="extended",
                                 yscrollcommand=vsb.set)
        vsb.configure(command=self.tree.yview)
        self.tree.heading("cod",    text="Cód. PDM")
        self.tree.heading("desc",   text="Descrição")
        self.tree.heading("status", text="Status")
        self.tree.column("cod",    width=90,   anchor="center", stretch=False)
        self.tree.column("desc",   width=9999, anchor="w",      stretch=True)
        self.tree.column("status", width=80,   anchor="center", stretch=False)
        self.tree.pack(fill="both", expand=True)

        # ── Card 3: Ações (rodapé) ────────────────────────────────────────────
        c3 = _card(parent, "3.  Ações")
        c3.pack(fill="x", padx=12, pady=(0,8))

        ar = ctk.CTkFrame(c3, fg_color="transparent")
        ar.pack(fill="x", padx=14, pady=(0,4))
        _btn(ar, "⚡  Buscar e Extrair", self._buscar_e_extrair,
             variant="primary").pack(side="left", padx=(0,8))
        self.btn_exp_cat = _btn(ar, "📥  Exportar CATMATs Encontrados",
                                self._exp_catmats, variant="ghost")
        self.btn_exp_cat.configure(state="disabled")
        self.btn_exp_cat.pack(side="left", padx=(0,16))
        self.lbl_exp_status = _lbl(ar, "", size=11, color=C_TEXT_MED)
        self.lbl_exp_status.pack(side="left", expand=True, fill="x")

        dr = ctk.CTkFrame(c3, fg_color="transparent")
        dr.pack(fill="x", padx=14, pady=(0,4))
        _lbl(dr, "Data de Início:", color=C_TEXT_MED).pack(side="left", padx=(0,6))
        self.var_ini2 = tk.StringVar()
        _entry(dr, textvariable=self.var_ini2, placeholder="DD-MM-AAAA", width=130)            .pack(side="left")
        _lbl(dr, "Data Final:", color=C_TEXT_MED).pack(side="left", padx=(16,6))
        self.var_fim2 = tk.StringVar()
        _entry(dr, textvariable=self.var_fim2, placeholder="DD-MM-AAAA", width=130)            .pack(side="left")

        cr = ctk.CTkFrame(c3, fg_color="transparent")
        cr.pack(fill="x", padx=14, pady=(0,10))
        self.btn_pb = _btn(cr, "⏸  Pausar Busca", self._pausar_busca,
                           variant="secondary", width=130)
        self.btn_pb.configure(state="disabled")
        self.btn_pb.pack(side="left", padx=(0,8))
        self.btn_cb = _btn(cr, "✖  Cancelar Busca", self._cancelar_busca,
                           variant="danger", width=130)
        self.btn_cb.configure(state="disabled")
        self.btn_cb.pack(side="left", padx=(0,16))
        self.btn_ini_exp = _btn(cr,
            "▶  Iniciar Extração com CATMATs Encontrados",
            self._iniciar_exp, variant="success")
        self.btn_ini_exp.configure(state="disabled")
        self.btn_ini_exp.pack(side="left")

        # ── LOG HELPERS ───────────────────────────────────────────────────────────
    def _log(self, msg: str, tag: str = ""):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.configure(state="disabled")

    def set_status(self, txt): self.lbl_status.configure(text=txt)
    def set_status_explorador(self, txt): self.lbl_exp_status.configure(text=txt)
    def _stat(self, key, val): self._stats[key].configure(text=str(val))



    # ── CALLBACKS ABA 1 ───────────────────────────────────────────────────────
    def _escolher_arquivo(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel/CSV","*.xlsx *.csv"),("Todos","*.*")])
        if p: self.var_arquivo.set(p)

    def _escolher_pasta(self):
        p = filedialog.askdirectory()
        if p: self.var_pasta.set(p)

    def _toggle_pasta(self):
        if self.var_salvar_corr.get():
            self.frame_pasta.pack(fill="x", padx=14, pady=3)
        else:
            self.frame_pasta.pack_forget()

    def _start(self):
        arq = self.var_arquivo.get().strip()
        if not arq:
            messagebox.showerror("Arquivo obrigatório",
                                 "Selecione um arquivo de códigos."); return
        try:
            df_c = pd.read_excel(arq) if arq.lower().endswith(".xlsx") \
                   else pd.read_csv(arq, sep=";")
            if "codigoItemCatalogo" not in df_c.columns:
                messagebox.showerror("Coluna ausente",
                    "O arquivo deve ter a coluna 'codigoItemCatalogo'."); return
            codigos = pd.Series(df_c["codigoItemCatalogo"]).dropna()\
                        .astype(int).drop_duplicates().tolist()
        except Exception as e:
            messagebox.showerror("Erro ao ler arquivo", str(e)); return
        d_i, d_f, err = validar_e_obter_datas(self.var_ini1.get(), self.var_fim1.get())
        if err: messagebox.showerror("Data inválida", err); return
        self._iniciar_processo(codigos, self.var_fmt.get(), d_i, d_f)

    def _cancelar(self):
        if not self.processing: return
        self.processing = False
        pausar_extracao.set()   # desbloqueia wait() na thread para ela poder sair

    def _pausar(self):
        if pausar_extracao.is_set():
            # Estava rodando → pausar
            pausar_extracao.clear()
            self.btn_pause.configure(text="▶  Retomar")
            self.set_status("Status: Pausado")
        else:
            # Estava pausado → retomar
            pausar_extracao.set()
            self.btn_pause.configure(text="⏸  Pausar")
            self.set_status("Status: Retomando…")

    def _salvar_log(self):
        p = filedialog.asksaveasfilename(defaultextension=".txt",
                                         filetypes=[("Texto","*.txt")])
        if p:
            try:
                with open(p,"w",encoding="utf-8") as f: f.write(self.log.get("1.0","end"))
                messagebox.showinfo("Salvo", f"Log salvo em:\n{p}")
            except Exception as e:
                messagebox.showerror("Erro", str(e))

    # ── CALLBACKS ABA 2 ───────────────────────────────────────────────────────
    # ── SPINNER (overlay translúcido) ─────────────────────────────────────────
    def _show_spinner(self, msg="Buscando…"):
        """Exibe overlay com spinner animado sobre a aba."""
        self._spinner_active = True
        self._spinner_frame = ctk.CTkFrame(self, fg_color="#FFFFFF",
                                           corner_radius=12,
                                           border_width=1, border_color=C_BORDER)
        self._spinner_frame.place(relx=0.5, rely=0.5, anchor="center")
        self._spinner_chars = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
        self._spinner_idx   = 0
        self._lbl_spin_icon = ctk.CTkLabel(self._spinner_frame,
            text=self._spinner_chars[0],
            font=("Segoe UI", 28), text_color=C_ACCENT)
        self._lbl_spin_icon.pack(padx=40, pady=(22,4))
        self._lbl_spin_msg = ctk.CTkLabel(self._spinner_frame,
            text=msg, font=("Segoe UI", 13), text_color=C_TEXT_MED)
        self._lbl_spin_msg.pack(padx=40, pady=(0,22))
        self._animate_spinner()

    def _animate_spinner(self):
        if not self._spinner_active: return
        self._spinner_idx = (self._spinner_idx + 1) % len(self._spinner_chars)
        self._lbl_spin_icon.configure(text=self._spinner_chars[self._spinner_idx])
        self.after(80, self._animate_spinner)

    def _hide_spinner(self):
        self._spinner_active = False
        if hasattr(self, "_spinner_frame") and self._spinner_frame.winfo_exists():
            self._spinner_frame.destroy()

    def _buscar_pdms(self, acao_pos_busca=None):
        entrada = self.var_classe.get().strip()
        if not entrada:
            messagebox.showerror("Campo vazio", "Informe ao menos um código de Classe."); return

        partes = [p.strip() for p in entrada.split(";") if p.strip()]
        invalidas = [p for p in partes if not p.isdigit()]
        if invalidas:
            messagebox.showerror("Código inválido",
                f"Valores não numéricos: {', '.join(invalidas)}\n"
                "Use apenas números separados por ;")
            return

        self._show_spinner(f"Buscando PDMs de {len(partes)} classe(s)…")
        self.lbl_pdm_count.configure(text="")

        def _thread():
            todos_dfs = []; erros = []
            for cod in partes:
                self.after(0, lambda c=cod: self._lbl_spin_msg.configure(
                    text=f"Buscando classe {c}…"))
                res = buscar_pdms_por_classe(int(cod), URL_BASE, TIMEOUT)
                if res is None:
                    erros.append(cod)
                else:
                    df_c, _ = res
                    todos_dfs.append(df_c)
            self.after(0, lambda: self._on_pdms_carregados(todos_dfs, erros, acao_pos_busca))

        threading.Thread(target=_thread, daemon=True).start()

    def _on_pdms_carregados(self, todos_dfs, erros, acao_pos_busca=None):
        self._hide_spinner()
        if not todos_dfs:
            self.lbl_pdm_count.configure(text="Nenhum PDM encontrado.")
            self.lista_pdms_df = pd.DataFrame(); self._fill_tree([]); return

        df = pd.concat(todos_dfs, ignore_index=True).drop_duplicates(subset=["codigoPdm"])
        df["_col_codigo"] = df["codigoPdm"]
        df["_col_desc"]   = df["nomePdm"]
        df["_col_status"] = df["statusPdm"]
        self.lista_pdms_df        = df
        self._todos_dfs_por_classe = todos_dfs  # preservar para mapa classe→catmats
        self._fill_tree([[r["_col_codigo"], r["_col_desc"], r["_col_status"]]
                         for _, r in df.iterrows()])

        msg = f"{len(df)} PDMs de {len(todos_dfs)} classe(s)"
        if erros: msg += f"  ·  ⚠ Falha: {', '.join(erros)}"
        self.lbl_pdm_count.configure(text=msg)
        if erros:
            messagebox.showwarning("Classes com falha",
                f"Não foi possível buscar: {', '.join(erros)}")
        self.var_filtro.set("todos")
        self.btn_exp_cat.configure(state="disabled")
        self.btn_ini_exp.configure(state="disabled")
        self.lista_catmats = []

        if acao_pos_busca == "extrair":
            self._continuar_busca_e_extrai(df)

    def _filtrar(self):
        if self.lista_pdms_df.empty: return
        f = self.var_filtro.get(); df = self.lista_pdms_df
        if f == "ativo":    df = df[df["_col_status"] == "Ativo"]
        elif f == "inativo": df = df[df["_col_status"] == "Inativo"]
        self._fill_tree([[r["_col_codigo"], r["_col_desc"], r["_col_status"]]
                         for _, r in df.iterrows()])
        # Manter checkbox em sincronia após filtrar
        if hasattr(self, "var_sel_todos"): self.var_sel_todos.set(False)

    def _toggle_selecionar_todos(self):
        """Seleciona ou deseleciona todos os itens visíveis na tabela."""
        items = self.tree.get_children()
        if self.var_sel_todos.get():
            self.tree.selection_set(items)
        else:
            self.tree.selection_remove(items)

    def _fill_tree(self, rows):
        for i in self.tree.get_children(): self.tree.delete(i)
        for row in rows:
            tag = "ativo" if str(row[2]).lower() == "ativo" else "inativo"
            self.tree.insert("", "end", values=row, tags=(tag,))
        self.tree.tag_configure("ativo",   foreground=C_GREEN)
        self.tree.tag_configure("inativo", foreground=C_TEXT_LIGHT)
        # Resetar checkbox ao recarregar
        if hasattr(self, "var_sel_todos"): self.var_sel_todos.set(False)

    def _exp_pdms(self):
        rows = [self.tree.item(i,"values") for i in self.tree.get_children()]
        if not rows: messagebox.showerror("Vazio","Nenhum PDM."); return
        p = filedialog.asksaveasfilename(defaultextension=".csv",
                                         initialfile="PDMs_exportados.csv",
                                         filetypes=[("CSV","*.csv")])
        if p:
            pd.DataFrame(rows, columns=["Código PDM","Descrição","Status"])\
              .to_csv(p, index=False, sep=";", encoding="utf-8-sig")
            messagebox.showinfo("Exportado", f"Salvo em:\n{p}")

    def _pdms_selecionados_codigos(self):
        """Retorna lista de codigoPdm dos itens selecionados na árvore."""
        sel = self.tree.selection()
        return [int(self.tree.item(i,"values")[0]) for i in sel] if sel else []

    def _pdms_sel(self):
        sel = self.tree.selection()
        return [int(self.tree.item(i,"values")[0]) for i in sel] if sel else []

    def _buscar_avulso(self):
        txt = self.txt_avulso.get("1.0", "end").strip()
        if not txt: messagebox.showerror("Vazio","Informe ao menos um código PDM."); return
        pdms = []; inv = []
        # Aceita separador ; ou nova linha
        for parte in txt.replace("\n", ";").split(";"):
            parte = parte.strip()
            if not parte: continue
            try: pdms.append(int(parte))
            except ValueError: inv.append(parte)
        if inv: messagebox.showwarning("Inválidos", f"Ignorados: {', '.join(inv)}")
        if pdms: self._start_busca(pdms, "apenas_buscar")

    def _buscar_catmats(self):
        pdms = self._pdms_sel()
        if not pdms: messagebox.showerror("Nenhum selecionado","Selecione PDMs."); return
        self._start_busca(pdms, "apenas_buscar")

    def _buscar_e_extrair(self):
        pdms = self._pdms_sel()
        if not pdms: messagebox.showerror("Nenhum selecionado","Selecione PDMs."); return
        self._start_busca(pdms, "extrair")

    def _start_busca(self, pdms, acao):
        global cancelar_busca_catmat
        cancelar_busca_catmat = False
        pausar_busca_catmat.set()
        self.btn_pb.configure(state="normal")
        self.btn_cb.configure(state="normal")
        self.btn_pb.configure(text="⏸  Pausar Busca")
        self.set_status_explorador("Iniciando busca…")
        threading.Thread(target=self._thread_busca,
                         args=(pdms, acao), daemon=True).start()

    def _thread_busca(self, pdms, acao):
        df1, err1 = buscar_catmats_por_pdm(pdms, URL_BASE, TIMEOUT, self)
        df2 = None; err2 = []
        if err1 and not cancelar_busca_catmat:
            self.after(0, lambda: self.set_status_explorador(
                f"2ª tentativa para {len(err1)} PDMs…"))
            time.sleep(2)
            df2, err2 = buscar_catmats_por_pdm(err1, URL_BASE, TIMEOUT, self)
        dfs = [d for d in [df1, df2] if d is not None and not d.empty]
        df_final = pd.concat(dfs, ignore_index=True) if dfs else None
        self.after(0, lambda: self._on_busca(df_final, err2, acao))

    def _on_busca(self, df, falhas, acao):
        self.btn_pb.configure(state="disabled")
        self.btn_cb.configure(state="disabled")
        if df is not None and "codigoItem" in df.columns:
            self.lista_catmats = df["codigoItem"].dropna().astype(int).tolist()
            n = len(self.lista_catmats)
            msg = f"✅ {n} CATMATs encontrados"
            if falhas: msg += f" · ⚠ {len(falhas)} PDMs com falha"
            self.set_status_explorador(msg)
            self.btn_exp_cat.configure(state="normal")
            self.btn_ini_exp.configure(state="normal", text=f"▶  Iniciar Extração com {n} CATMATs Encontrados")
            if falhas:
                messagebox.showwarning("PDMs com falha",
                    f"Falha persistente em:\n{', '.join(map(str,falhas))}")
            if acao == "extrair": self._iniciar_exp()
        else:
            self.lista_catmats = []
            if not cancelar_busca_catmat:
                self.set_status_explorador("Nenhum CATMAT encontrado.")
            self.btn_exp_cat.configure(state="disabled")
            self.btn_ini_exp.configure(state="disabled")

    def _exp_catmats(self):
        if not self.lista_catmats:
            messagebox.showerror("Vazio","Nenhum CATMAT."); return
        p = filedialog.asksaveasfilename(defaultextension=".csv",
                                         initialfile="CATMATs_descobertos.csv",
                                         filetypes=[("CSV","*.csv")])
        if p:
            pd.DataFrame(self.lista_catmats, columns=["codigoItemCatalogo"])\
              .to_csv(p, index=False, sep=";")
            messagebox.showinfo("Exportado", f"{len(self.lista_catmats)} CATMATs:\n{p}")

    def _pausar_busca(self):
        if pausar_busca_catmat.is_set():
            pausar_busca_catmat.clear()
            self.btn_pb.configure(text="▶  Retomar Busca")
            self.set_status_explorador("Busca pausada.")
        else:
            pausar_busca_catmat.set()
            self.btn_pb.configure(text="⏸  Pausar Busca")
            self.set_status_explorador("Retomando…")

    def _toggle_pasta_por_classe(self):
        """Mostra/oculta o campo de pasta quando o checkbox é marcado."""
        if self.var_arquivo_por_classe.get():
            self.frame_pasta_classes.pack(fill="x", padx=0, pady=(6,0))
        else:
            self.frame_pasta_classes.pack_forget()
            self.var_pasta_classes.set("")

    def _escolher_pasta_classes(self):
        p = filedialog.askdirectory(title="Escolha a pasta de destino dos arquivos por classe")
        if p:
            self.var_pasta_classes.set(p)

    def _cancelar_busca(self):
        global cancelar_busca_catmat
        cancelar_busca_catmat = True

    def _iniciar_exp(self):
        if not self.lista_catmats:
            messagebox.showerror("Vazio","Nenhum CATMAT disponível."); return
        d_i, d_f, err = validar_e_obter_datas(self.var_ini2.get(), self.var_fim2.get())
        if err: messagebox.showerror("Data inválida", err); return
        self._log(f"🔎 {len(self.lista_catmats)} CATMATs via explorador.", "info")
        por_classe = self.var_arquivo_por_classe.get()
        # Usa o mapa já construído em _on_busca_e_extrai (se existir)
        mapa = getattr(self, "_catmats_por_classe", {}) if por_classe else {}
        if por_classe and not mapa:
            messagebox.showwarning("Aviso",
                "Use Buscar e Extrair para gerar arquivos por classe.")
            return
        self._iniciar_processo(self.lista_catmats, self.var_fmt.get(), d_i, d_f,
                               catmats_por_classe=mapa)
        self.after(100, lambda: self.tabs.set("  Extração por CATMAT  "))

    def _buscar_e_extrair_classes(self):
        """Fluxo automatizado: para cada classe, faz PDMs → CATMATs → Extração → Salva."""
        entrada = self.var_classe.get().strip()
        if not entrada:
            messagebox.showerror("Campo vazio", "Informe ao menos um código de Classe."); return

        partes = [p.strip() for p in entrada.split(";") if p.strip()]
        invalidas = [p for p in partes if not p.isdigit()]
        if invalidas:
            messagebox.showerror("Código inválido",
                "Valores nao numericos: " + ", ".join(invalidas)); return

        por_classe = self.var_arquivo_por_classe.get()
        pasta_dest = self.var_pasta_classes.get().strip() if por_classe else ""

        if por_classe and len(partes) > 1 and not pasta_dest:
            messagebox.showerror("Pasta obrigatória",
                "Selecione uma pasta de destino para salvar os arquivos por classe."); return

        d_i, d_f, err = validar_e_obter_datas(self.var_ini2.get(), self.var_fim2.get())
        if err: messagebox.showerror("Data inválida", err); return

        fmt = self.var_fmt.get()
        self._pasta_classes_destino = pasta_dest
        self.processing   = True
        self.total_baixados = 0
        self.count_corrigidas = 0
        self.count_vazios = 0
        pausar_extracao.set()
        pausar_busca_catmat.set()  # necessário para o fluxo automatizado

        self.log.configure(state="normal"); self.log.delete("1.0","end")
        self.log.configure(state="disabled")
        for k, v in [("k_proc","0"),("k_reg","0"),("k_corr","0"),("k_vaz","0")]:
            self._stat(k, v)
        self.progress.set(0); self.lbl_pct.configure(text="0%")
        self.set_status("Status: Iniciando…")
        self.btn_start.configure(state="disabled")
        self.btn_cancel.configure(state="normal", fg_color=C_RED,
                                   hover_color="#992B1E", text_color=C_SURFACE)
        self.btn_pause.configure(state="normal", text="⏸  Pausar")
        self.btn_log.configure(state="disabled")

        self._log("💾 Formato: " + ("CSV" if fmt == "csv" else "Excel"), "info")
        if d_i or d_f:
            def fd(s):
                p = s.split("-"); return p[2]+"-"+p[1]+"-"+p[0]
            txt = "📅 Filtro de datas:"
            if d_i: txt += "  Início: " + fd(d_i)
            if d_f: txt += "  Fim: "    + fd(d_f)
            self._log(txt, "date")
        self._log("📋 " + str(len(partes)) + " classe(s): " + " | ".join(partes) + "\n", "info")
        if pasta_dest:
            self._log("📂 Destino: " + pasta_dest, "info")

        self.after(100, lambda: self.tabs.set("  Extração por CATMAT  "))
        threading.Thread(
            target=self._fluxo_classes_thread,
            args=(partes, pasta_dest, fmt, d_i, d_f),
            daemon=True
        ).start()

    def _fluxo_classes_thread(self, classes_lista, pasta_dest, fmt, d_ini, d_fim):
        """
        Thread principal do fluxo por classe.
        Para cada classe: PDMs → CATMATs → Registros de Preços → salva arquivo + relatório
        Retry em todos os níveis: classes, PDMs e CATMATs com erro.
        """
        ext          = "csv" if fmt == "csv" else "xlsx"
        total_classes_orig = len(classes_lista)
        salvar_corr  = self.var_salvar_corr.get()
        pasta_corr   = self.var_pasta.get()
        total_catmats_acum = 0
        classes_com_falha  = []   # classes que não retornaram PDMs

        # ── Helper: extrai os registros de preços de um CATMAT ────────────────
        # Retorna (baixados, tipo_resultado)
        # tipo_resultado: "ok" | "vazio" | "erro"
        def _extrair_catmat(codigo, writer, reg_baixados, reg_esperados, pag_corrompidas):
            baixados = 0; pagina_atual = 1; total_paginas = None
            try:
                _, csv_text = ler_pagina_catmat(codigo, 1, URL_BASE, 500, TIMEOUT, d_ini, d_fim)
                if csv_text and csv_text.startswith("ERRO_CONEXAO"):
                    return 0, "conexao"
                if csv_text is None or csv_text.startswith("ERRO_REQUISICAO"):
                    return 0, "erro"   # erro de API — pode ser retentado

                m = re.search(r"totalRegistros\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                reg_esp = int(m.group(1)) if m else 0
                reg_esperados[codigo] = reg_esp
                if reg_esp == 0:
                    return 0, "vazio"  # genuinamente sem dados — não retenta

                while True:
                    if not self.processing: break
                    is_c, csv_c = pagina_corrompida(csv_text)
                    df_pag = parse_csv_text(csv_c)
                    if is_c:
                        pag_corrompidas.setdefault(codigo, []).append(str(pagina_atual))
                        if salvar_corr and pasta_corr:
                            dest_c = os.path.join(pasta_corr,
                                "cod_" + str(codigo) + "_pag_" + str(pagina_atual) + "_corr.csv")
                            try:
                                with open(dest_c, "w", encoding="utf-8-sig") as f:
                                    f.write(csv_text)
                            except Exception: pass
                        self._ui(lambda cod=codigo, p=pagina_atual:
                            self._log("⚠️  Cód " + str(cod) + " Pág " + str(p) + ": corrigida.", "warn"))
                        self.count_corrigidas += 1
                        self._ui(lambda v=self.count_corrigidas: self._stat("k_corr", v))
                    else:
                        self._ui(lambda cod=codigo, p=pagina_atual:
                            self._log("✅  Cód " + str(cod) + " Pág " + str(p) + ": OK.", "ok"))

                    if df_pag is not None and not df_pag.empty:
                        df_pag.loc[:, "codigoItemCatalogo"] = str(codigo)
                        df_proc = processar_dataframe_final(df_pag, ordem_final_colunas)
                        baixados += len(df_proc)
                        self.total_baixados += len(df_proc)
                        writer.write_dataframe(df_proc)
                        reg = self.total_baixados
                        self._ui(lambda r=reg:
                            self._stat("k_reg", str(r).replace(",",".")))

                    if total_paginas is None:
                        mp = re.search(r"total\s*p[áa]ginas?\s*:\s*(\d+)", csv_text, re.IGNORECASE)
                        total_paginas = int(mp.group(1)) if mp else 1
                    pagina_atual += 1
                    if pagina_atual > total_paginas: break
                    time.sleep(0.5)
                    _, csv_text = ler_pagina_catmat(codigo, pagina_atual, URL_BASE,
                                                    500, TIMEOUT, d_ini, d_fim)
                    if csv_text is None or csv_text.startswith("ERRO_"): break

                return baixados, "ok"
            except Exception as e:
                etxt = "❌  Erro CATMAT " + str(codigo) + ": " + str(e)
                self._ui(lambda t=etxt: self._log(t, "err"))
                return 0, "erro"

        # ── Helper: processa uma classe completa ──────────────────────────────
        def _processar_classe(classe, idx_c, total_c):
            nonlocal total_catmats_acum
            ic, tc = idx_c, total_c
            sep = "─" * 50
            hdr = sep + "\n📦  CLASSE " + classe + "  (" + str(ic) + "/" + str(tc) + ")\n" + sep
            self._ui(lambda h=hdr: self._log("\n" + h, "date"))
            self._ui(lambda c=classe, i=ic, t=tc:
                self.set_status("Status: Classe " + c + " (" + str(i) + "/" + str(t) + ") — PDMs…"))

            # ── 1. PDMs ───────────────────────────────────────────────────────
            resultado = buscar_pdms_por_classe(int(classe), URL_BASE, TIMEOUT)
            if resultado is None:
                return False   # falhou — será retentada
            df_pdms, _ = resultado
            pdms_lista  = df_pdms["codigoPdm"].astype(int).tolist()
            n_pdms = len(pdms_lista)
            self._ui(lambda c=classe, n=n_pdms:
                self._log("✅  Classe " + c + ": " + str(n) + " PDMs.", "ok"))

            # ── 2. CATMATs ────────────────────────────────────────────────────
            self._ui(lambda c=classe, n=n_pdms:
                self.set_status("Status: Classe " + c + " — CATMATs (" + str(n) + " PDMs)…"))

            _log_pdm = lambda m: self._log("  🔍 " + m, "info")
            df_cat1, erros1 = buscar_catmats_por_pdm(pdms_lista, URL_BASE, TIMEOUT, self, log_fn=_log_pdm)
            df_cat2 = None; erros2 = []
            if erros1 and not cancelar_busca_catmat:
                n_e1 = len(erros1)
                self._ui(lambda n=n_e1:
                    self._log("♻️  2ª tentativa (10s) para " + str(n) + " PDMs com erro…", "warn"))
                time.sleep(3)
                df_cat2, erros2 = buscar_catmats_por_pdm(erros1, URL_BASE, TIMEOUT, self, log_fn=_log_pdm)
            # 3ª tentativa para PDMs que ainda falharam
            df_cat3 = None; erros3 = []
            if erros2 and not cancelar_busca_catmat:
                n_e2 = len(erros2)
                self._ui(lambda n=n_e2:
                    self._log("♻️  3ª tentativa (20s) para " + str(n) + " PDMs com erro…", "warn"))
                time.sleep(8)
                df_cat3, erros3 = buscar_catmats_por_pdm(erros2, URL_BASE, TIMEOUT, self, log_fn=_log_pdm)
                if erros3:
                    falha_pdm = ", ".join(map(str, erros3))
                    self._ui(lambda t=falha_pdm:
                        self._log("❌  PDMs sem resposta após 3 tentativas: " + t, "err"))

            dfs_c = [d for d in [df_cat1, df_cat2, df_cat3] if d is not None and not d.empty]
            df_catmats = pd.concat(dfs_c, ignore_index=True) if dfs_c else None
            if df_catmats is None or "codigoItem" not in df_catmats.columns:
                self._ui(lambda c=classe:
                    self._log("⚠️  Classe " + c + ": nenhum CATMAT. Pulando.", "warn"))
                return True   # PDMs foram encontrados mas sem CATMATs — não é falha de PDMs

            catmats_lista = df_catmats["codigoItem"].dropna().astype(int).tolist()
            n_cat = len(catmats_lista)
            self._ui(lambda c=classe, n=n_cat:
                self._log("✅  Classe " + c + ": " + str(n) + " CATMATs.", "ok"))

            # ── 3. Extração dos Registros de Preços ───────────────────────────
            nome_arq = "classe_" + classe + "." + ext
            caminho  = os.path.join(pasta_dest, nome_arq) if pasta_dest else nome_arq
            writer   = CSVChunkWriter(caminho) if fmt == "csv" else ExcelChunkWriter(caminho)

            reg_baixados    = {}
            reg_esperados   = {}
            pag_corrompidas = {}
            total_baixados_classe = 0
            vazios_classe         = 0
            catmats_com_erro      = []
            total_cat  = len(catmats_lista)
            writer_lock = threading.Lock()
            state_lock  = threading.Lock()
            comp_count  = [0]

            def _processar_resultado(codigo, dfs_e_meta, tipo, reg_esp, pag_corr):
                nonlocal total_baixados_classe, vazios_classe, total_catmats_acum
                if tipo == "conexao":
                    self._ui(lambda: messagebox.showerror("Conexão",
                        "Erro de conexão com a API. Verifique sua rede."))
                    self.processing = False
                    return "conexao"
                elif tipo == "erro":
                    with state_lock:
                        catmats_com_erro.append(codigo)
                    etxt = "⚠️  CATMAT " + str(codigo) + ": erro na API, será retentado."
                    self._ui(lambda t=etxt: self._log(t, "warn"))
                elif tipo == "vazio":
                    vtxt = "ℹ️  " + str(codigo) + ": 0 registros."
                    self._ui(lambda t=vtxt: self._log(t, "info"))
                    with state_lock:
                        vazios_classe += 1
                        self.count_vazios += 1
                        v = self.count_vazios
                    self._ui(lambda vv=v: self._stat("k_vaz", vv))
                else:
                    baixados = sum(len(df) for df, _, _ in dfs_e_meta)
                    with writer_lock:
                        for df_proc, _, _ in dfs_e_meta:
                            writer.write_dataframe(df_proc)
                    with state_lock:
                        total_baixados_classe += baixados
                        reg_baixados[codigo]   = baixados
                        reg_esperados[codigo]  = reg_esp
                        self.total_baixados   += baixados
                        for k, v in pag_corr.items():
                            pag_corrompidas.setdefault(k, []).extend(v)
                        n_corr = sum(len(v) for v in pag_corr.values())
                        self.count_corrigidas += n_corr
                        reg = self.total_baixados
                    self._ui(lambda r=reg: self._stat("k_reg", f"{r:,}".replace(",",".")))
                    for _, is_c, pag in dfs_e_meta:
                        if is_c:
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"⚠️  Cód {cod} Pág {p}: corrigida.", "warn"))
                    if pag_corr:
                        with state_lock:
                            cv = self.count_corrigidas
                        self._ui(lambda vv=cv: self._stat("k_corr", vv))
                with state_lock:
                    comp_count[0] += 1
                    total_catmats_acum += 1
                    comp = comp_count[0]
                    tca  = total_catmats_acum
                self._ui(lambda v=tca: self._stat("k_proc", v))
                pct = (idx_c - 1 + comp / total_cat) / total_c
                self._ui(lambda c=classe, i=comp, t=total_cat, p=pct:
                    (self.set_status("Status: Classe " + c + " — " + str(i) + "/" + str(t)),
                     self.progress.set(p),
                     self.lbl_pct.configure(text=str(int(p*100)) + "%")))
                return tipo

            # Extração paralela (4 workers)
            with ThreadPoolExecutor(max_workers=1) as executor:
                futures = {
                    executor.submit(_fetch_catmat_registros, cod,
                                    d_ini, d_fim, salvar_corr, pasta_corr): cod
                    for cod in catmats_lista
                }
                for future in as_completed(futures):
                    pausar_extracao.wait()
                    if not self.processing:
                        executor.shutdown(wait=False, cancel_futures=True); break
                    cod, dfs_m, tipo, reg_e, pag_c = future.result()
                    res = _processar_resultado(cod, dfs_m, tipo, reg_e, pag_c)
                    if res == "conexao": break

            # Retry paralelo dos CATMATs com erro
            for espera in [15, 30]:
                if not catmats_com_erro or not self.processing: break
                n_err = len(catmats_com_erro)
                self._ui(lambda n=n_err, e=espera:
                    self._log("♻️  Retry " + str(n) + " CATMAT(s) com erro (aguardando " +
                              str(e) + "s)…", "warn"))
                time.sleep(espera)
                retry_list = list(catmats_com_erro); catmats_com_erro.clear()
                with ThreadPoolExecutor(max_workers=1) as executor:
                    futures = {
                        executor.submit(_fetch_catmat_registros, cod,
                                        d_ini, d_fim, salvar_corr, pasta_corr): cod
                        for cod in retry_list
                    }
                    for future in as_completed(futures):
                        if not self.processing: break
                        cod, dfs_m, tipo, reg_e, pag_c = future.result()
                        _processar_resultado(cod, dfs_m, tipo, reg_e, pag_c)

            if catmats_com_erro:
                n_def = len(catmats_com_erro)
                self._ui(lambda n=n_def:
                    self._log("❌  " + str(n) + " CATMAT(s) sem resposta após 3 tentativas.", "err"))

            if not self.processing: return True

            # ── 4. Finalizar arquivo desta classe ─────────────────────────────
            parts = writer.finalize()
            arqs  = ", ".join(os.path.basename(p) for p in parts) if parts else "(sem dados)"
            self._ui(lambda c=classe, a=arqs, n=total_baixados_classe:
                self._log("📁  Classe " + c + " — " + str(n) + " registros → " + a, "info"))

            # ── 5. Relatório de integridade desta classe ──────────────────────
            rel_nome    = "Relatorio_Integridade_" + classe + ".xlsx"
            rel_caminho = os.path.join(pasta_dest, rel_nome) if pasta_dest else rel_nome
            try:
                wb = Workbook(); ws = wb.active; ws.title = "Integridade_" + classe
                ws.append(["codigoItemCatalogo","esperados","baixados","paginas","status"])
                for c in catmats_lista:
                    bx = int(reg_baixados.get(c, 0))
                    ex = int(reg_esperados.get(c, 0))
                    pg = pag_corrompidas.get(c, [])
                    d  = abs(ex - bx)
                    st = ("OK" if d == 0 else
                          "OK (divergencia: " + str(bx) + "/" + str(ex) + ")" if d <= 2 else
                          "Inconsistencia Grave (" + str(bx) + "/" + str(ex) + ")")
                    ws.append([c, ex, bx, ", ".join(map(str, pg)), st])
                if catmats_com_erro:
                    ws.append([])
                    ws.append(["--- CATMATs sem resposta apos 3 tentativas ---"])
                    for c in catmats_com_erro:
                        ws.append([c, 0, 0, "", "ERRO_API_PERSISTENTE"])
                wb.save(rel_caminho)
                self._ui(lambda r=rel_nome:
                    self._log("📊  Relatório: " + r, "info"))
            except Exception as e:
                etxt = "⚠️ Relatório classe " + classe + " não salvo: " + str(e)
                self._ui(lambda t=etxt: self._log(t, "warn"))
            return True

        # ── Loop principal por classe ─────────────────────────────────────────
        for idx_classe, classe in enumerate(classes_lista, 1):
            if not self.processing: break
            ok = _processar_classe(classe, idx_classe, total_classes_orig)
            if not ok:
                self._ui(lambda c=classe:
                    self._log("⚠️  Classe " + c + ": sem PDMs na 1ª tentativa. Fila de retry.", "warn"))
                classes_com_falha.append(classe)

        # ── Retry classes sem PDMs (10s → 20s) ───────────────────────────────
        for num_tent, espera in enumerate([3, 8], 2):
            if not classes_com_falha or not self.processing: break
            n_f = len(classes_com_falha)
            self._ui(lambda n=n_f, e=espera, t=num_tent:
                self._log("\n♻️  " + str(n) + " classe(s) sem PDMs — tentativa " +
                          str(t) + "/3 (aguardando " + str(e) + "s)…", "warn"))
            time.sleep(espera)
            ainda_falha = []
            for idx, classe in enumerate(classes_com_falha, 1):
                if not self.processing: break
                ok = _processar_classe(classe, idx, len(classes_com_falha))
                if not ok:
                    ainda_falha.append(classe)
            classes_com_falha = ainda_falha

        if classes_com_falha:
            falhas = ", ".join(classes_com_falha)
            self._ui(lambda t=falhas:
                self._log("❌  Classes sem PDMs após 3 tentativas: " + t, "err"))

        # ── Todas as classes processadas ──────────────────────────────────────
        self._ui(self._finalizar_fluxo_classes)

    def _finalizar_fluxo_classes(self):
        """Chamado ao término de todas as classes."""
        foi_cancelado = not self.processing
        self.processing = False
        self.progress.set(1.0); self.lbl_pct.configure(text="100%")
        self.set_status("Status: Concluído!" if not foi_cancelado else "Status: Cancelado")
        self._log(
            "\n🎉 Todas as classes processadas com sucesso!" if not foi_cancelado
            else "\n🛑 Extração cancelada.", "info")

        self.btn_start.configure(state="normal")
        self.btn_cancel.configure(state="disabled",
                                   fg_color="#E4E7EF", hover_color=C_BORDER,
                                   text_color=C_TEXT)
        self.btn_pause.configure(state="disabled", text="⏸  Pausar")
        self.btn_log.configure(state="normal")

        pasta = getattr(self, "_pasta_classes_destino", "").strip()
        if pasta:
            messagebox.showinfo("Concluído",
                "Extração finalizada!\nTodos os arquivos foram salvos em:\n" + pasta)
        else:
            messagebox.showinfo("Concluído", "Extração finalizada!")

        # ── MOTOR DE EXTRAÇÃO ─────────────────────────────────────────────────────
    def _iniciar_processo(self, codigos, fmt, d_ini, d_fim, catmats_por_classe=None):
        if not codigos: return
        self.processing            = True
        self.codigos_lista         = codigos
        self._data_inicio          = d_ini
        self._data_fim             = d_fim
        self._fmt                  = fmt
        self._catmats_por_classe_ativo = catmats_por_classe or {}
        # Pasta escolhida pelo usuário para salvar os arquivos por classe
        self._pasta_classes_destino = self.var_pasta_classes.get().strip()             if hasattr(self, "var_pasta_classes") else ""
        self.paginas_corrompidas   = {}
        self.registros_esperados   = {}
        self.registros_baixados    = {}
        self.total_baixados        = 0
        self.count_corrigidas      = 0
        self.count_vazios          = 0
        pausar_extracao.set()

        # Se há arquivo por classe, usamos um writer por classe (criados sob demanda)
        # Senão, writer único
        if self._catmats_por_classe_ativo:
            self.writer = None  # será None; usamos self._writers_por_classe
            self._writers_por_classe = {}  # classe → writer
        else:
            self.writer = CSVChunkWriter("dados_completos_extraidos.csv") \
                          if fmt == "csv" else \
                          ExcelChunkWriter("dados_completos_extraidos.xlsx")
            self._writers_por_classe = {}

        self.log.configure(state="normal"); self.log.delete("1.0","end")
        self.log.configure(state="disabled")
        self._log(f"💾 Formato: {'CSV' if fmt == 'csv' else 'Excel'}", "info")
        if d_ini or d_fim:
            def fd(s): p = s.split("-"); return f"{p[2]}-{p[1]}-{p[0]}"
            txt = "📅 Filtro de datas:"
            if d_ini: txt += f"  Início: {fd(d_ini)}"
            if d_fim:  txt += f"  Fim: {fd(d_fim)}"
            self._log(txt, "date")
        self._log(f"🔎 {len(codigos)} códigos carregados.\n", "info")

        for k, v in [("k_proc",f"0 / {len(codigos)}"),
                     ("k_reg","0"),("k_corr","0"),("k_vaz","0")]:
            self._stat(k, v)
        self.progress.set(0); self.lbl_pct.configure(text="0%")
        self.set_status("Status: Processando…")
        self.btn_start.configure(state="disabled")
        self.btn_cancel.configure(state="normal", fg_color=C_RED,
                                   hover_color="#992B1E", text_color=C_SURFACE)
        self.btn_pause.configure(state="normal", text="⏸  Pausar")
        self.btn_log.configure(state="disabled")
        # Lança extração em thread separada — UI continua responsiva
        self._extracao_thread_obj = threading.Thread(
            target=self._extracao_thread, daemon=True)
        self._extracao_thread_obj.start()

    # ─────────────────────────────────────────────────────────────────────────
    # MOTOR DE EXTRAÇÃO — roda 100% em thread separada
    # Comunicação com UI exclusivamente via self.after(0, callback)
    # ─────────────────────────────────────────────────────────────────────────

    def _get_writer_para(self, codigo: int):
        if not self._catmats_por_classe_ativo:
            return self.writer
        classe_do_cod = "outras"
        for classe, cats in self._catmats_por_classe_ativo.items():
            if codigo in cats:
                classe_do_cod = classe; break
        if classe_do_cod not in self._writers_por_classe:
            ext   = "csv" if self._fmt == "csv" else "xlsx"
            # Salva direto na pasta escolhida pelo usuário (se informada)
            pasta = getattr(self, "_pasta_classes_destino", "").strip()
            nome_arq = f"classe_{classe_do_cod}.{ext}"
            caminho  = os.path.join(pasta, nome_arq) if pasta else nome_arq
            self._writers_por_classe[classe_do_cod] = (
                CSVChunkWriter(caminho) if self._fmt == "csv"
                else ExcelChunkWriter(caminho))
        return self._writers_por_classe[classe_do_cod]

    def _ui(self, fn):
        """Agenda fn() na thread principal de forma segura."""
        self.after(0, fn)

    def _extracao_thread(self):
        """Thread de extração paralela — 4 CATMATs simultâneos."""
        codigos     = self.codigos_lista
        total       = len(codigos)
        salvar_corr = self.var_salvar_corr.get()
        pasta_corr  = self.var_pasta.get()
        d_ini       = self._data_inicio
        d_fim       = self._data_fim
        writer_locks: dict = {}   # id(writer) → Lock
        state_lock  = threading.Lock()
        comp_count  = [0]

        def _wlock(codigo):
            w = self._get_writer_para(codigo)
            k = id(w)
            if k not in writer_locks:
                writer_locks[k] = threading.Lock()
            return w, writer_locks[k]

        with ThreadPoolExecutor(max_workers=1) as executor:
            futures = {
                executor.submit(_fetch_catmat_registros, cod,
                                d_ini, d_fim, salvar_corr, pasta_corr): cod
                for cod in codigos
            }
            for future in as_completed(futures):
                pausar_extracao.wait()
                if not self.processing:
                    executor.shutdown(wait=False, cancel_futures=True); break

                codigo, dfs_e_meta, tipo, reg_esp, pag_corr = future.result()

                with state_lock:
                    comp_count[0] += 1
                    comp = comp_count[0]

                if tipo == "conexao":
                    self._ui(lambda: messagebox.showerror("Conexão",
                        "Erro de conexão com a API. Verifique sua rede."))
                    self.processing = False; break

                elif tipo in ("erro", "vazio"):
                    txt = (f"ℹ️  {codigo}: sem registro (erro API)." if tipo == "erro"
                           else f"ℹ️  {codigo}: 0 registros.")
                    with state_lock:
                        self.count_vazios += 1
                        v = self.count_vazios
                        self.registros_baixados[codigo]  = 0
                        self.registros_esperados[codigo] = reg_esp
                    self._ui(lambda t=txt: self._log(t, "info"))
                    self._ui(lambda vv=v: self._stat("k_vaz", vv))

                else:  # "ok"
                    baixados = sum(len(df) for df, _, _ in dfs_e_meta)
                    w, wlock = _wlock(codigo)
                    with wlock:
                        for df_proc, _, _ in dfs_e_meta:
                            w.write_dataframe(df_proc)
                    with state_lock:
                        self.total_baixados             += baixados
                        self.registros_baixados[codigo]  = baixados
                        self.registros_esperados[codigo] = reg_esp
                        reg = self.total_baixados
                        for k, v in pag_corr.items():
                            self.paginas_corrompidas.setdefault(k, []).extend(v)
                        n_c = sum(len(v) for v in pag_corr.values())
                        self.count_corrigidas += n_c
                        cv = self.count_corrigidas
                    self._ui(lambda r=reg: self._stat("k_reg", f"{r:,}".replace(",",".")))
                    for _, is_c, pag in dfs_e_meta:
                        if is_c:
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"⚠️  Cód {cod} Pág {p}: corrigida.", "warn"))
                        else:
                            self._ui(lambda cod=codigo, p=pag:
                                self._log(f"✅  Cód {cod} Pág {p}: OK.", "ok"))
                    if pag_corr:
                        self._ui(lambda vv=cv: self._stat("k_corr", vv))

                pct = comp / total
                self._ui(lambda p=pct, c=comp, t=total: (
                    self.progress.set(p),
                    self.lbl_pct.configure(text=f"{int(p*100)}%"),
                    self.set_status(f"Status: Processando... ({c}/{t})"),
                    self._stat("k_proc", f"{c} / {t}")
                ))

        self._ui(self._finalizar)

    def _finalizar(self):
        """Chamado na thread principal via after() ao término da extração."""
        foi_cancelado = not self.processing
        self.processing = False
        self.progress.set(1.0); self.lbl_pct.configure(text="100%")
        self.set_status("Status: Concluído!" if not foi_cancelado else "Status: Cancelado")
        self._log("\n🎉 Extração concluída!" if not foi_cancelado
                  else "\n🛑 Extração cancelada.", "info")

        # Finalizar writers
        parts = []
        if self._catmats_por_classe_ativo and self._writers_por_classe:
            for classe, w in self._writers_por_classe.items():
                p = w.finalize(); parts.extend(p)
                if p: self._log(f"📂 Classe {classe}: {', '.join(p)}", "info")
        elif self.writer:
            parts = self.writer.finalize()

        if parts:
            self._log(f"💾 Arquivos gerados: {', '.join(parts)}", "info")

        # Relatório de integridade
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Relatorio Integridade"
            ws.append(["codigoItemCatalogo","esperados","baixados","paginas","status"])
            for c in self.codigos_lista:
                bx = int(self.registros_baixados.get(c,0))
                ex = int(self.registros_esperados.get(c,0))
                pg = self.paginas_corrompidas.get(c,[])
                d  = abs(ex-bx)
                st = ("OK" if d==0 else
                      f"OK (divergencia: {bx}/{ex})" if d<=2 else
                      f"Inconsistencia Grave ({bx}/{ex})")
                ws.append([c,ex,bx,", ".join(map(str,pg)),st])
            wb.save("Relatorio_Integridade.xlsx")
            self._log("📊 Relatorio_Integridade.xlsx gerado.", "info")
        except Exception as e:
            self._log(f"⚠️ Relatório não salvo: {e}", "warn")

        self.btn_start.configure(state="normal")
        self.btn_cancel.configure(state="disabled",
                                   fg_color="#E4E7EF", hover_color=C_BORDER,
                                   text_color=C_TEXT)
        self.btn_pause.configure(state="disabled", text="⏸  Pausar")
        self.btn_log.configure(state="normal")

        if not parts:
            messagebox.showinfo("Sem dados","Nenhum dado válido baixado.")
            return

        n_classes = len(self._writers_por_classe) if self._catmats_por_classe_ativo else 0
        ext = os.path.splitext(parts[0])[1]

        resumo_linhas = [
            "Processo Concluido!",
            chr(8212)*40,
            f"Codigos Processados:     {len(self.codigos_lista)}",
            f"Registros Consolidados:  {self.total_baixados:,}",
            f"Paginas Corrigidas:      {self.count_corrigidas}",
            f"Codigos sem Registros:   {self.count_vazios}",
        ]
        if n_classes > 1:
            resumo_linhas.append(f"Arquivos por classe:     {n_classes}")
        messagebox.showinfo("Resumo", "\n".join(resumo_linhas))

        if n_classes > 1:
            pasta_dest = getattr(self, "_pasta_classes_destino", "").strip()
            nomes = "\n".join(os.path.basename(p) for p in parts)
            if pasta_dest:
                # Arquivos já foram escritos direto na pasta pelo writer
                # Apenas copiar o relatório de integridade para a mesma pasta
                try:
                    shutil.copy("Relatorio_Integridade.xlsx",
                                os.path.join(pasta_dest, "Relatorio_Integridade.xlsx"))
                except Exception:
                    pass
                self._log(f"📁 {len(parts)} arquivo(s) em: {pasta_dest}", "info")
                messagebox.showinfo("Concluído",
                    f"{len(parts)} arquivo(s) salvos em:\n{pasta_dest}\n\n{nomes}")
            else:
                # Sem pasta definida: pede agora e copia
                pasta_dest = filedialog.askdirectory(
                    title=f"Escolha a pasta para salvar os {len(parts)} arquivo(s)")
                if pasta_dest:
                    for arq in parts:
                        shutil.copy(arq, os.path.join(pasta_dest, os.path.basename(arq)))
                    try:
                        shutil.copy("Relatorio_Integridade.xlsx",
                                    os.path.join(pasta_dest, "Relatorio_Integridade.xlsx"))
                    except Exception:
                        pass
                    self._log(f"📁 {len(parts)} arquivo(s) salvos em: {pasta_dest}", "info")
                    messagebox.showinfo("Concluído",
                        f"{len(parts)} arquivo(s) salvos em:\n{pasta_dest}\n\n{nomes}")
                else:
                    messagebox.showwarning("Atenção",
                        f"Nenhuma pasta escolhida. Arquivos na pasta do programa:\n{nomes}")
        else:
            ultimo = parts[-1]
            tipos  = [("Excel","*.xlsx")] if ext == ".xlsx" else [("CSV","*.csv")]
            dest   = filedialog.asksaveasfilename(
                        defaultextension=ext,
                        initialfile=os.path.basename(ultimo),
                        filetypes=tipos)
            if dest:
                if not dest.lower().endswith(ext): dest += ext
                shutil.copy(ultimo, dest)
                messagebox.showinfo("Salvo",
                    f"Dados salvos em:\n{dest}\n\nRelatorio de integridade na pasta do programa.")
            else:
                messagebox.showwarning("Atencao", f"Arquivo permanece em:\n{ultimo}")

# =============================================================================
# PALETA  —  neutros Gov.br + acento azul Gov + verde BPS + amarelo BPS
# =============================================================================
C_BG         = "#F4F5F7"   # cinza-papel (fundo geral)
C_SURFACE    = "#FFFFFF"   # branco (cards / frames)
C_BORDER     = "#DDE1E9"   # borda sutil
C_TEXT       = "#1A1D23"   # quase-preto
C_TEXT_MED   = "#555B6E"   # texto secundário
C_TEXT_LIGHT = "#8A92A6"   # placeholder / hint
C_ACCENT     = "#1351B4"   # azul Gov.br (primário)
C_ACCENT_H   = "#0C3784"   # hover do azul
C_GREEN      = "#168821"   # verde BPS (sucesso)
C_GREEN_H    = "#0E5C17"   # hover verde
C_YELLOW     = "#FFCD07"   # amarelo BPS (destaque / faixa)
C_ORANGE     = "#E37222"   # aviso
C_RED        = "#C0392B"   # erro / cancelar
C_LOG_BG     = "#13141A"   # terminal escuro
C_LOG_FG     = "#E8EAF0"   # texto terminal

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# =============================================================================
# LÓGICA DE NEGÓCIO
# =============================================================================

pausar_extracao     = threading.Event()
pausar_busca_catmat = threading.Event()
cancelar_busca_catmat = False

requests.packages.urllib3.disable_warnings(
    requests.packages.urllib3.exceptions.InsecureRequestWarning
)

# Session compartilhada — reutiliza conexões TCP/TLS entre todas as requisições
# Evita o overhead de handshake (~200-400ms) a cada chamada
_http = requests.Session()
_http.verify = False
_http.headers.update({"Accept-Encoding": "gzip, deflate", "Connection": "keep-alive"})
_adapter = requests.adapters.HTTPAdapter(
    pool_connections=4, pool_maxsize=12, max_retries=0
)
_http.mount("https://", _adapter)
_http.mount("http://",  _adapter)

URL_BASE = "https://dadosabertos.compras.gov.br"
TIMEOUT  = 120

ordem_final_colunas = [
    "idCompra","idItemCompra","forma","modalidade","criterioJulgamento",
    "numeroItemCompra","descricaoItem","codigoItemCatalogo","nomeUnidadeFornecimento",
    "siglaUnidadeFornecimento","nomeUnidadeMedida","capacidadeUnidadeFornecimento","siglaUnidadeMedida",
    "Unidade de Fornecimento","capacidade","quantidade","precoUnitario","Preco Total","percentualMaiorDesconto",
    "niFornecedor","nomeFornecedor","marca","codigoUasg","nomeUasg",
    "codigoMunicipio","municipio","estado","codigoOrgao","nomeOrgao",
    "poder","esfera","dataCompra","dataHoraAtualizacaoCompra","dataHoraAtualizacaoItem",
    "dataResultado","dataHoraAtualizacaoUasg","codigoClasse","nomeClasse",
]


class ExcelChunkWriter:
    def __init__(self, base_filename, sheet_name="Dados CATMAT", max_rows_per_file=1_000_000):
        self.base_filename = base_filename
        self.sheet_name    = sheet_name
        self.max_rows      = max_rows_per_file
        self.part          = 1
        self.header: List[str] = []
        self.current_row_count = 0
        self.files_saved: List[str] = []
        self._new_workbook()

    def _filepath(self):
        base, ext = os.path.splitext(self.base_filename)
        if not ext or ext.lower() != ".xlsx": ext = ".xlsx"
        return f"{base}_part{self.part}{ext}"

    def _new_workbook(self):
        self.wb = Workbook(); self.ws = self.wb.active
        self.ws.title = self.sheet_name
        self.header_written = False; self.current_row_count = 0

    def _ensure_header(self, columns):
        if not self.header: self.header = list(columns)
        if not self.header_written:
            self.ws.append(self.header); self.header_written = True

    def _rollover_if_needed(self):
        if self.current_row_count + 1 > self.max_rows:
            path = self._filepath(); self.wb.save(path); self.files_saved.append(path)
            self.part += 1; self._new_workbook()
            if self.header: self.ws.append(self.header); self.header_written = True

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        self._ensure_header(list(df.columns))
        for col in self.header:
            if col not in df.columns: df[col] = pd.NA
        df = df[self.header]
        for _, row in df.iterrows():
            self._rollover_if_needed()
            self.ws.append([None if pd.isna(v) else v for v in row])
            self.current_row_count += 1

    def finalize(self) -> List[str]:
        if self.header_written and self.current_row_count > 0:
            path = self._filepath(); self.wb.save(path)
            if path not in self.files_saved: self.files_saved.append(path)
        return self.files_saved


class CSVChunkWriter:
    def __init__(self, base_filename, sep=";", encoding="utf-8-sig", max_rows_per_file=1_000_000):
        self.base_filename = base_filename; self.sep = sep
        self.encoding = encoding; self.max_rows = max_rows_per_file
        self.part = 1; self.current_row_count = 0
        self.files_saved: List[str] = []; self.header_written = False

    def _filepath(self):
        base, ext = os.path.splitext(self.base_filename)
        if not ext or ext.lower() != ".csv": ext = ".csv"
        return f"{base}_part{self.part}{ext}"

    def write_dataframe(self, df: pd.DataFrame):
        if df is None or df.empty: return
        if self.current_row_count + len(df) > self.max_rows:
            self.part += 1; self.current_row_count = 0; self.header_written = False
        path = self._filepath()
        df.to_csv(path, sep=self.sep, index=False,
                  mode="a" if self.header_written else "w",
                  header=not self.header_written, encoding=self.encoding)
        self.header_written = True; self.current_row_count += len(df)
        if path not in self.files_saved: self.files_saved.append(path)

    def finalize(self) -> List[str]:
        return self.files_saved


def converter_data_para_api(data_dd_mm_yyyy: str) -> Optional[str]:
    s = data_dd_mm_yyyy.strip()
    if not s: return None
    try:
        p = s.split("-")
        if len(p) != 3: return None
        dd, mm, yyyy = p
        if len(dd) == 2 and len(mm) == 2 and len(yyyy) == 4:
            int(dd); int(mm); int(yyyy)
            return f"{yyyy}-{mm}-{dd}"
    except (ValueError, AttributeError):
        pass
    return None


def validar_e_obter_datas(ini: str, fim: str):
    i_api = f_api = None
    if ini.strip():
        i_api = converter_data_para_api(ini)
        if i_api is None:
            return None, None, f"Data de Inicio invalida: '{ini}'\nUse DD-MM-AAAA (ex: 01-01-2024)"
    if fim.strip():
        f_api = converter_data_para_api(fim)
        if f_api is None:
            return None, None, f"Data Final invalida: '{fim}'\nUse DD-MM-AAAA (ex: 31-12-2024)"
    return i_api, f_api, None


def parse_csv_text(csv_text: str) -> pd.DataFrame:
    lines = [ln for ln in csv_text.splitlines() if ln.strip()]
    if not lines: return pd.DataFrame()
    try:
        return pd.read_csv(StringIO("\n".join(lines)), sep=";", dtype=str,
                           engine="python", on_bad_lines="warn", quoting=3)
    except Exception:
        return pd.DataFrame()


def ler_pagina_catmat(codigo, pagina, URL_BASE, TAMANHO_PAGINA, TIMEOUT,
                      data_compra_inicio=None, data_compra_fim=None):
    URL = f"{URL_BASE}/modulo-pesquisa-preco/1.1_consultarMaterial_CSV"
    params = {"tamanhoPagina": TAMANHO_PAGINA, "codigoItemCatalogo": int(codigo), "pagina": int(pagina)}
    if data_compra_inicio: params["dataCompraInicio"] = data_compra_inicio
    if data_compra_fim:    params["dataCompraFim"]    = data_compra_fim
    tentativas = 0
    while tentativas < 2:
        try:
            resp = _http.get(URL, params=params, timeout=TIMEOUT)
            if resp.status_code == 429:
                time.sleep(15 if tentativas == 0 else 30); tentativas += 1; continue
            resp.raise_for_status()
            return None, resp.content.decode("utf-8-sig", errors="replace")
        except requests.exceptions.ConnectionError as e:
            return None, f"ERRO_CONEXAO: {e}"
        except requests.exceptions.RequestException as e:
            return None, f"ERRO_REQUISICAO: {e}"
    return None, f"ERRO_REQUISICAO: 429 persistente para CATMAT {codigo}"


def _normalizar_campo(item: dict, *candidatos, default=""):
    """Retorna o primeiro campo encontrado no dict entre os candidatos."""
    for c in candidatos:
        if c in item and item[c] is not None:
            return item[c]
    return default


def buscar_pdms_por_classe(codigo_classe: int, URL_BASE: str, TIMEOUT: int):
    URL = f"{URL_BASE}/modulo-material/3_consultarPdmMaterial"
    all_pdms = []; pagina_atual = 1; total_paginas = 1; total_registros_api = 0
    TAMANHO_PAGINA = 500
    while pagina_atual <= total_paginas:
        try:
            resp = _http.get(URL, params={
                "codigoClasse": codigo_classe, "pagina": pagina_atual,
                "tamanhoPagina": TAMANHO_PAGINA, "bps": "false"
            }, timeout=TIMEOUT)
            resp.raise_for_status(); data = resp.json()
            if "resultado" in data:
                # Logar as keys do primeiro item para diagnóstico
                if pagina_atual == 1 and data["resultado"]:
                    print(f"DEBUG keys PDM: {list(data['resultado'][0].keys())}")
                    print(f"DEBUG primeiro item: {data['resultado'][0]}")
                all_pdms.extend(data["resultado"])
            if pagina_atual == 1:
                total_registros_api = int(data.get("totalRegistros", 0))
                total_paginas = math.ceil(total_registros_api / TAMANHO_PAGINA) if total_registros_api > 0 else 1
                print(f"DEBUG: Registros: {total_registros_api} | Paginas: {total_paginas}")
            pagina_atual += 1; time.sleep(0.5)
        except Exception as e:
            print(f"DEBUG erro busca classe {codigo_classe}: {e}")
            return None
    if not all_pdms: return None

    # Normalizar campos — a API pode retornar nomes variados
    rows_norm = []
    for item in all_pdms:
        cod  = _normalizar_campo(item, "codigoPdm", "codigo", "id", "codigoItem")
        desc = _normalizar_campo(item, "nomePdm", "nome", "descricao", "descricaoPdm", "descricaoItem")
        # status pode ser bool True/False, string "ATIVO"/"INATIVO", ou inteiro
        raw_status = _normalizar_campo(item, "statusPdm", "status", "ativo", "situacao")
        if isinstance(raw_status, bool):
            status = "Ativo" if raw_status else "Inativo"
        elif isinstance(raw_status, str):
            status = "Ativo" if raw_status.upper() in ("ATIVO", "TRUE", "S", "SIM", "1") else "Inativo"
        elif isinstance(raw_status, (int, float)):
            status = "Ativo" if raw_status == 1 else "Inativo"
        else:
            status = "Ativo"
        rows_norm.append({"codigoPdm": cod, "nomePdm": desc, "statusPdm": status,
                          "_classe": str(codigo_classe)})

    df = pd.DataFrame(rows_norm).drop_duplicates(subset=["codigoPdm"])
    return df, total_registros_api


def buscar_catmats_por_pdm(codigos_pdm, URL_BASE, TIMEOUT, app,
                           log_fn=None, max_workers=5):
    """
    Busca CATMATs de múltiplos PDMs em paralelo usando ThreadPoolExecutor.
    max_workers=5 → ~5x mais rápido sem sobrecarregar a API.
    """
    global cancelar_busca_catmat
    URL   = f"{URL_BASE}/modulo-material/4_consultarItemMaterial"
    total = len(codigos_pdm)
    all_catmats  = []
    pdms_com_erro = []
    completed     = 0
    lock          = threading.Lock()   # protege all_catmats e pdms_com_erro

    def _fetch_pdm(idx_pdm):
        """Worker: busca todas as páginas de CATMATs de um PDM."""
        i, pdm_code = idx_pdm
        if cancelar_busca_catmat:
            return
        resultados = []; pagina_atual = 1; total_paginas = 1
        try:
            while pagina_atual <= total_paginas:
                if cancelar_busca_catmat: break
                resp = _http.get(URL, params={
                    "codigoPdm": pdm_code, "pagina": pagina_atual,
                    "tamanhoPagina": 500, "bps": "false"
                }, timeout=TIMEOUT)
                if resp.status_code == 429:
                    time.sleep(15); continue          # rate-limit: aguarda e repete
                resp.raise_for_status()
                data = resp.json()
                if "resultado" in data:
                    resultados.extend(data["resultado"])
                if pagina_atual == 1:
                    total_paginas = data.get("totalPaginas", 1)
                pagina_atual += 1
                if pagina_atual <= total_paginas:
                    time.sleep(0.2)                   # throttle entre páginas do mesmo PDM
        except Exception:
            with lock:
                pdms_com_erro.append(pdm_code)
            return
        if resultados:
            with lock:
                all_catmats.extend(resultados)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_fetch_pdm, (i, pdm)): (i, pdm)
            for i, pdm in enumerate(codigos_pdm)
        }
        for future in as_completed(futures):
            pausar_busca_catmat.wait()        # respeita pausa
            if cancelar_busca_catmat:
                app.after(0, lambda: app.set_status_explorador("Busca cancelada."))
                executor.shutdown(wait=False, cancel_futures=True)
                break
            completed += 1
            i, pdm_code = futures[future]
            msg = f"PDM {pdm_code} ({completed}/{total})..."
            app.after(0, lambda m=msg: app.set_status_explorador(m))
            if log_fn and (completed % 5 == 0 or completed == total):
                app.after(0, lambda m=msg: log_fn(m))

    return pd.DataFrame(all_catmats) if all_catmats else None, pdms_com_erro


def pagina_corrompida(csv_text: str):
    if not csv_text: return False, csv_text
    linhas = [ln for ln in csv_text.splitlines() if ln.strip()]
    if not linhas: return False, csv_text
    try:
        hi = next(i for i, ln in enumerate(linhas)
                  if not ln.lower().startswith(("totalregistros:", "totalpaginas:")))
        header_line = linhas[hi]; ncols = len(header_line.split(";"))
    except StopIteration:
        return False, csv_text
    if ncols == 0: return False, csv_text
    out = list(linhas[:hi]) + [header_line]
    buf = ""; corrigido = False
    for ln in linhas[hi+1:]:
        if ln.lower().startswith(("totalregistros:", "totalpaginas:")):
            if buf: out.append(buf); buf = ""
            out.append(ln); continue
        atual = buf + ln.replace("\r","").replace("\n","")
        if len(atual.split(";")) < ncols:
            buf = atual + " "; corrigido = True
        else:
            out.append(atual); buf = ""
    if buf: out.append(buf)
    return corrigido, "\n".join(out)


def processar_dataframe_final(df: pd.DataFrame, ordem_colunas: List[str]) -> pd.DataFrame:
    if df.empty: return df
    fc = df.columns[0]
    df = df[~df[fc].astype(str).str.contains("totalRegistros|totalPaginas",
                                              case=False, na=False)].copy()
    if df.empty: return df

    def uf(row):
        # Usa as 4 colunas na ordem definida; inclui só as que têm valor
        partes = []
        for k in ["nomeUnidadeFornecimento", "siglaUnidadeFornecimento",
                  "capacidadeUnidadeFornecimento", "siglaUnidadeMedida"]:
            v = row.get(k)
            if pd.notna(v) and str(v).strip() and str(v).strip() not in ("nan", "None"):
                partes.append(str(v).strip())
        return " ".join(partes)

    df["Unidade de Fornecimento"] = df.apply(uf, axis=1)

    def tof(v):
        if pd.isna(v): return 0.0
        try: return float(str(v).replace(".", "").replace(",", "."))
        except: return 0.0

    df["Preco Total"] = df["precoUnitario"].apply(tof) * df["quantidade"].apply(tof)
    for col in ["nomeUnidadeMedida","percentualMaiorDesconto"]:
        if col in df.columns and (df[col].isnull().all() or
                                   df[col].astype(str).str.strip().eq("").all()):
            df = df.drop(columns=[col])
    exist = [c for c in ordem_colunas if c in df.columns]
    extra = [c for c in df.columns if c not in exist]
    return df[exist + extra]


# =============================================================================
# COMPONENTES DE UI  (helpers)
# =============================================================================

WELCOME = """\
Olá! Bem-vindo ao Extrator de CATMATs Pro.

Sua ferramenta para extrair e descobrir dados no Portal de Compras Governamentais!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
O que este programa faz?
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Este programa possui duas funcoes principais em abas separadas:

  1. Extracao por CATMAT (esta aba)
     Se voce ja tem uma lista de codigos de materiais (CATMATs), esta aba
     busca todas as informacoes de compras, corrige problemas nos dados e
     consolida tudo em um arquivo Excel ou CSV.

  2. Extracao por Classes (aba ao lado)
     Se voce quer descobrir novos itens, pode comecar com o codigo de uma
     ou mais Classes, encontrar todos os Padroes Descritivos de Materiais
     (PDMs) dentro delas e, em seguida, listar todos os CATMATs relacionados
     para extracao.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Primeiros Passos
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  - Para uma extracao direta com uma lista pronta, use esta aba.
    O arquivo Excel ou CSV deve ter a coluna: codigoItemCatalogo

  - Para descobrir itens, use a aba "Extracao por Classes" e, ao final,
    envie os CATMATs encontrados para a extracao nesta aba.

  - Utilize os filtros de data (DD-MM-AAAA) para restringir os resultados
    a um periodo especifico de compras (Data de Inicio e Data Final).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Acompanhe todo o processo em tempo real neste log. Bom trabalho!
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""


def _lbl(parent, text, size=12, weight="normal", color=C_TEXT, **kw):
    return ctk.CTkLabel(parent, text=text, font=("Segoe UI", size, weight),
                        text_color=color, **kw)


def _btn(parent, text, command, variant="secondary", width=0, **kw):
    pal = {
        "primary":   (C_SURFACE,  C_ACCENT,  C_SURFACE,  C_ACCENT_H),
        "success":   (C_SURFACE,  C_GREEN,   C_SURFACE,  C_GREEN_H),
        "danger":    (C_SURFACE,  C_RED,     C_SURFACE,  "#992B1E"),
        "secondary": (C_TEXT,     "#E4E7EF", C_TEXT,     C_BORDER),
        "ghost":     (C_ACCENT,   "transparent", C_ACCENT_H, "#E8EDF8"),
    }
    tc, bg, htc, hbg = pal.get(variant, pal["secondary"])
    return ctk.CTkButton(parent, text=text, command=command,
                         font=("Segoe UI", 12), fg_color=bg, text_color=tc,
                         hover_color=hbg, corner_radius=6,
                         width=width, height=32, **kw)


def _entry(parent, textvariable=None, placeholder="", width=200, **kw):
    return ctk.CTkEntry(parent, textvariable=textvariable,
                        placeholder_text=placeholder,
                        font=("Segoe UI", 12),
                        fg_color=C_SURFACE, text_color=C_TEXT,
                        border_color=C_BORDER, border_width=1,
                        corner_radius=6, width=width,
                        placeholder_text_color=C_TEXT_LIGHT, **kw)


def _sep(parent, pady=(6,6)):
    ctk.CTkFrame(parent, height=1, fg_color=C_BORDER,
                 corner_radius=0).pack(fill="x", padx=14, pady=pady)


def _card(parent, title="", **kw):
    outer = ctk.CTkFrame(parent, fg_color=C_SURFACE, corner_radius=8,
                         border_width=1, border_color=C_BORDER, **kw)
    if title:
        _lbl(outer, title, size=11, weight="bold", color=C_TEXT_MED)\
            .pack(anchor="w", padx=14, pady=(10,4))
        _sep(outer, pady=(0,6))
    return outer


# =============================================================================
# APLICATIVO PRINCIPAL


# =============================================================================
if __name__ == "__main__":
    app = App()
    app.mainloop()